import MinClass
from openpyxl import Workbook, load_workbook


def read_cell_i_lista_by_rad(filnavn, start_sheet, start_rad):
    # Åpne filen
    wb = load_workbook(filnavn, data_only=True)
    ws = wb.worksheets[start_sheet]

    # loope gjennom hver rad fra start_rad til siste rad
    for i in range(start_rad, ws.max_row + 1):
        # Kopier raden til en variabel
        row = ws[i]
        # Sjekk om alle cellene i raden er tomme
        if all(cell.value is None for cell in row):
            # Hvis alle cellene er tomme, gå til neste rad
            continue
        # Lag en liste med verdiene i hver celle i raden
        rad = [cell.value for cell in row]
        # Returner listen
        yield rad


def finn_sheet_etter_verdi(filnavn, start_sheet, objekt, verdi):
    start_rad = 7
    # Åpne filen
    wb = load_workbook(filnavn, data_only=True)
    ws = wb.worksheets[start_sheet]

    for ws in wb.worksheets:
        celle = str(ws["C"+str(1)].value)
        if celle == objekt:
            rad = sjekk_tom_rad(start_rad, ws)
#            sjekk = input(f"Celle: {celle}rad: {rad}")
            if rad > 0:
                ws["AB"+str(start_rad)].value = verdi
                start_rad += 1
                wb.save(filnavn)

            return ws.title
            break


def snu(liste):
    try:
        status = liste
        status_list = [str(s) for s in status]
        status_list = status.split(",")
        r_list = status_list[::-1]

    except AttributeError:
        # Håndter feilen ved å gi status_list en default verdi
        status_list = []
        print("Feil: status er ikke en string og kan ikke splittes.")

    return r_list


def sjekk_tom_rad(start_rad, ws):
    logikk = True
    while logikk:
        row = ws[start_rad]
        if all(cell.value is None for cell in row[3:4]):
            # Hvis alle cellene er tomme, gå til neste rad
            rad = start_rad
            logikk = False
        else:
            start_rad = start_rad + 1
    return rad


def read_nord():
    try:
        # Åpne objektlista
        wb = load_workbook("NORD_OBJ.xlsx", data_only=True)
    except FileNotFoundError:
        print("Filen NORD_OBJ.xlsx ble ikke funnet")
        return
    except PermissionError:
        print("Du har ikke tilgang til filen NORD_OBJ.xlsx")
        return

    for ws in wb.worksheets:
        rad = str(ws["C"+str(1)].value)
        yield rad
