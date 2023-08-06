"""
Data Deciling Module
"""

import pandas as pd
import numpy as np
from IPython.core.display import HTML
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl as op
import os
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN

class resultsData:
    """Class to generate deciles, actual vs predicted check, PSI check"""

    def genRank(self,y,y_pred):
        """
        Create the HTML styled table of the dataframe

        Parameters
        ----------
        y : Series
                Target column
        y_pred : Series
                Target Predicted values

        Returns
        -------
        rank_df : dataframe
                Ranked Dataframe
        """
        print("\n Ranking the population...")
        rank_df = pd.DataFrame({'tgt':y,'score':y_pred}).reset_index(drop=True)
        rank_df['Rank'] = rank_df.score.rank(method='dense',ascending=False).astype(int)
        rank_df['Rank Order'] = rank_df.score.rank(method='first',ascending=False).astype(int)
        return rank_df

    def decilesData(self,data,target,score,divisions=[],regions=[]):
        """
        Create the decile table in HTML style

        Parameters
        ----------
        data : dataframe
                Input dataframe.
        target : str
                Target column name
        score : str
                Score column
        divisions : List of str, optional
                List of divisions to generate the deciles
        regions : List of str, optional
                List of regions to generate the deciles

        Returns
        -------
        agg1 : dataframe
            Deciles Dataframe.
        """
        print("\nGenerating the dataframe fields for deciling...")
        data['nontarget'] = 1 - data[target]
        data['bucket'] = pd.qcut(data[score], 10)
        grouped = data.groupby('bucket', as_index = False)
        print('****check****')
        agg1 = pd.DataFrame()
        agg1['TARGET'] = grouped.sum()[target]
        agg1['NONTARGET'] = grouped.sum()['nontarget']
        agg1['TOTAL'] = agg1['TARGET'] + agg1['NONTARGET']
        agg1['PCT_TAR'] = grouped.mean()[target]*100
        decile_color = 'skyblue'

        print("\nDeciling starts...")
        
        if len(divisions) == 0 and len(regions) == 0:
            decile_by=[]
        else:
            decile_by=['DIVISION','REGION']
        if len(decile_by) > 1:
            print("\nGenereating the KS-Values/Spread...")
            agg1['CUM_TAR'] = grouped.sum()[target].groupby(level=[0]).cumsum()
            agg1['CUM_NONTAR'] = grouped.sum()['nontarget'].groupby(level=[0]).cumsum()
            agg1['DIST_TAR'] = agg1['CUM_TAR']/agg1['TARGET'].groupby(level=[0]).sum()*100
            agg1['DIST_NONTAR'] = agg1['CUM_NONTAR']/agg1['NONTARGET'].groupby(level=[0]).sum()*100
            agg1['SPREAD'] = (agg1['DIST_TAR'] - agg1['DIST_NONTAR'])
            d = {}
            print("\nDeciling based on Division/Region...")
            if agg1.index.names[0] == 'DIVISION':
                loop = divisions
            elif agg1.index.names[0] == 'REGION':
                loop = regions

            for i in loop:
                x = loop.index(i)
                d[x] = agg1.ix[i]
                d[x] = self.decile_labels(d[x],i,color=decile_color)
                d[x] = self.plot_pandas_style(d[x])
            print("\nDeciling ends...")
            return(d)
        else:
            print("\nGenereating the KS-Values/Spread...")
            agg1['CUM_TAR'] = grouped.sum()[target].cumsum()
            agg1['CUM_NONTAR'] = grouped.sum()['nontarget'].cumsum()
            agg1['DIST_TAR'] = agg1['CUM_TAR']/agg1['TARGET'].sum()*100
            agg1['DIST_NONTAR'] = agg1['CUM_NONTAR']/agg1['NONTARGET'].sum()*100
            agg1['SPREAD'] = (agg1['DIST_TAR'] - agg1['DIST_NONTAR'])
            print("\nDeciling ends...")
            return agg1

    def plots(self,data,target):
        """
        Plotting the graphs

        Parameters
        ----------
        data : dataframe
                Input dataframe.
        target : str
                Target column name

        Returns
        -------
        None
        """


        print("\nPlotting the Actual vs Predicted Chart...")
        plt.figure(figsize = (8, 10))
        plt.plot(data['DECILE'],data['ACTUAL'],label='Actual')
        plt.plot(data['DECILE'],data['PRED'],label='Pred')
        plt.xticks(range(10,110,10))
        plt.grid(True)
        plt.title('Actual vs Predicted', fontsize=20)
        plt.xlabel("Population %",fontsize=15)
        plt.ylabel(str(target) + " " + " %",fontsize=15)

        newpath = r'gains_lifts'
        if not os.path.exists(newpath):
            os.makedirs(newpath)
            plt.savefig(newpath+"/actual_vs_pred.png")
        else:
            plt.savefig(newpath+"/actual_vs_pred.png")

        xfile = op.load_workbook('documentation_final.xlsx')
        sheet = xfile['8. Actual vs Predicted']
        img = op.drawing.image.Image(newpath+"/actual_vs_pred.png")
        img.anchor = 'J7'
        sheet.add_image(img)
        xfile.save('documentation_final.xlsx')

        print("\nPlotting the Gains Chart...")
        plt.figure(figsize = (8, 10))
        X = data['DECILE'].tolist()
        X.append(0)
        Y = data['DIST_TAR'].tolist()
        Y.append(0)
        plt.plot(sorted(X),sorted(Y))
        plt.plot([0, 100], [0, 100],'r--')
        plt.xticks(range(0,110,10))
        plt.yticks(range(0,110,10))
        plt.grid(True)
        plt.title('Gains Chart', fontsize=20)
        plt.xlabel("Population %",fontsize=15)
        plt.ylabel(str(target) + str(" DISTRIBUTION") + " %",fontsize=15)

        if not os.path.exists(newpath):
            os.makedirs(newpath)
            plt.savefig(newpath+"/gains.png")
        else:
            plt.savefig(newpath+"/gains.png")

        xfile = op.load_workbook('documentation_final.xlsx')
        sheet = xfile['8. Actual vs Predicted']
        img = op.drawing.image.Image(newpath+"/gains.png")
        img.anchor = 'J60'
        sheet.add_image(img)
        xfile.save('documentation_final.xlsx')

        print("\nPlotting the Lift Chart...")
        plt.figure(figsize = (8, 10))
        plt.plot(data['DECILE'],data['LIFT'])
        plt.xticks(range(10,110,10))
        plt.grid(True)
        plt.title('Lift Chart', fontsize=20)
        plt.xlabel("Population %",fontsize=15)
        plt.ylabel("Lift",fontsize=15)
        if not os.path.exists(newpath):
            os.makedirs(newpath)
            plt.savefig(newpath+"/lifts.png")
        else:
            plt.savefig(newpath+"/lifts.png")

        xfile = op.load_workbook('documentation_final.xlsx')
        sheet = xfile['8. Actual vs Predicted']
        img = op.drawing.image.Image(newpath+"/lifts.png")
        img.anchor = 'J115'
        sheet.add_image(img)
        xfile.save('documentation_final.xlsx')

        plt.tight_layout()

    def reg_plots(self,data,target,divisions,regions,reg_split):
        """
        Plotting the graphs for the regions

        Parameters
        ----------
        data : dataframe
                Input dataframe.
        target : str
                Target column name
        divisions: List of str
                List of divisions
        regions : List of str
                List of regions
        reg_split : List of tuple
                List of positions where the regions belong to divisions

        Returns
        -------
        None
        """

        plt.figure(1,figsize=(20, 5))
        print("\nPlotting the Gains chart for the regions within the divisions...")

        for j in reg_split:
            for i in regions[j[0]-1:j[1]]:
                reg_df = data[data['REGION'] == i]
                X = reg_df['DECILE'].tolist()
                X.append(0)
                Y = reg_df['DIST_TAR'].tolist()
                Y.append(0)
                plt.plot(sorted(X),sorted(Y),label= i)
            plt.plot([0, 100], [0, 100],'r--')
            plt.xticks(range(0,110,10))
            plt.yticks(range(0,110,10))
            plt.legend(fontsize=12)
            plt.grid(True)
            plt.title('Gains Chart - '+divisions[0], fontsize=15)
            plt.xlabel("Population %",fontsize=15)
            plt.ylabel(str(target) + str(" DISTRIBUTION") + " %",fontsize=15)

        plt.tight_layout()
        print("\nPlotting for regions completed...")
        plt.savefig("gains_reg.png")

    def dec_plots(agg1,target,divisions):
        """
        Plotting the graphs for divisions

        Parameters
        ----------
        data : dataframe
                Input dataframe.
        target : str
                Target column name
        divisions : List of str
                List of divisions

        Returns
        -------
        None
        """
        plt.figure(1,figsize=(20, 5))

        print("\nPlotting the Actual vs Predicted Chart for the Divisions...")
        plt.subplot(131)
        for i in divisions:
            div_df = agg1[agg1['DIVISION'] == i]
            plt.plot(div_df['DECILE'],div_df['ACTUAL'],label=i + ' - Actual')
            plt.plot(div_df['DECILE'],div_df['PRED'],label=i + ' - Pred')
        plt.xticks(range(10,110,10))
        plt.legend(fontsize=12)
        plt.grid(True)
        plt.title('Actual vs Predicted', fontsize=20)
        plt.xlabel("Population %",fontsize=15)
        plt.ylabel(str(target) + " " + " %",fontsize=15)

        print("\nPlotting the Gains Chart for the Divisions...")
        plt.subplot(132)
        for i in divisions:
            X = i['DECILE'].tolist()
            X.append(0)
            Y = i['DIST_TAR'].tolist()
            Y.append(0)
            plt.plot(sorted(X),sorted(Y),label= i.name)
        plt.xticks(range(0,110,10))
        plt.plot([0, 100], [0, 100],'r--')
        plt.yticks(range(0,110,10))
        plt.legend(fontsize=12)
        plt.grid(True)
        plt.title('Gains Chart', fontsize=20)
        plt.xlabel("Population %",fontsize=15)
        plt.ylabel(str(target) + str(" DISTRIBUTION") + " %",fontsize=15)

        print("\nPlotting the Lift chart for the Divisions...")
        plt.subplot(133)
        for i in divisions:
            plt.plot(i['DECILE'],i['LIFT'], label = i.name)
        plt.xticks(range(10,110,10))
        plt.legend(fontsize=12)
        plt.grid(True)
        plt.title('Lift Chart', fontsize=20)
        plt.xlabel("Population %",fontsize=15)
        plt.ylabel("Lift",fontsize=15)

        plt.tight_layout()

        plt.savefig('gain_lift_divisions.png')

    def gains(self,data,target,score,divisions=[],regions=[],reg_split=[]):
        """
        Calculate and Plot the Actual vs Predicted values, Gains Chart, Lift Chart

        Parameters
        ----------
        data : dataframe
                Input dataframe.
        target : str
                Target column name
        score : str
                Score column
        divisions : List of str, optional
                List of divisions to generate the deciles
        regions : List of str, optional
                List of regions to generate the deciles
        reg_split : List of tuple
                List of positions where the regions belong to divisions

        Returns
        -------
        None
        """
        print("\n")
        data['bucket'] = pd.qcut(data['Rank Order'], 10)
        multiplier = 10
        grouped = data.groupby('bucket')

        print("\nGenerating the Actual vs Predicted values...")
        agg1 = pd.DataFrame()
        agg1['ACTUAL'] = grouped.mean()[target]*100
        agg1['PRED'] = grouped.mean()[score]*100
        if len(divisions) == 0 and len(regions) == 0:
            decile_by = []
        else:
            decile_by=['DIVISION','REGION']

        if len(decile_by) > 1:
            print("\nActual vs Predicted, Gains and Lift Plotting for Division/Region...")
            agg1['DIST_TAR'] = grouped.sum()[target].groupby(level=[0]).cumsum()/grouped.sum()[target].groupby(level=[0]).sum()*100
            agg1 = agg1.reset_index()
            if agg1.columns[0:1].item() == 'REGION':
                agg1=agg1.rename(columns = {agg1.columns[1:2].item():'DECILE'})
            elif agg1.columns[0:1].item() == 'DIVISION':
                agg1=agg1.rename(columns = {agg1.columns[1:2].item():'DECILE'})

            print("\nCalculating the Lift...")
            agg1['DECILE'] = agg1['DECILE']*multiplier
            agg1['LIFT'] = agg1['DIST_TAR']/agg1['DECILE']

            print("\nPlotting the charts for Divsion/Region...")
            if agg1.columns[0:1].item() == 'REGION':
                self.reg_plots(agg1,target,divisions,regions,reg_split)
            elif agg1.columns[0:1].item() == 'DIVISION':
                self.dec_plots(agg1,target,divisions)
        else:
            print("\nActual vs Predicted, Gains and Lift Plotting...")
            agg1['DIST_TAR'] = grouped.sum()[target].cumsum()/grouped.sum()[target].sum()*100
            agg1['DECILE'] = [1,2,3,4,5,6,7,8,9,10]
            agg1.index.name = 'BINS'
            agg1 = agg1.reset_index()

            print("\nCalculating the Lift...")
            agg1['DECILE'] = agg1['DECILE']*multiplier
            agg1['LIFT'] = agg1['DIST_TAR']/agg1['DECILE']

            print(agg1)

            print("\nWriting actual vs predicted to excel starts...")
            df = agg1[['DECILE','BINS','ACTUAL','PRED','DIST_TAR','LIFT']]
            df2 = agg1[['ACTUAL','PRED']]
            with pd.ExcelWriter("documentation_final.xlsx",mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name='8. Actual vs Predicted',startrow=4,startcol=0,index=False,header=False)
                df2.to_excel(writer, sheet_name='9. Model Stability',startrow=6,startcol=1,index=False,header=False)
            print("\nWriting actual vs predicted to excel ends...")
            print("\nPlotting the charts")
            self.plots(agg1,target)

    def calculate_psi(self,expected, actual, buckettype='bins', buckets=10, axis=0):
        '''Calculate the PSI (population stability index) across all variables

        Parameters:
        ----------
        expected: numpy matrix of original values (Training)
        actual: numpy matrix of new values, same size as expected (Validation)
        buckettype: type of strategy for creating buckets, bins splits into even splits, quantiles splits into quantile buckets
        buckets: number of quantiles to use in bucketing variables
        axis: axis by which variables are defined, 0 for vertical, 1 for horizontal

        Returns:
        ----------
        psi_values: ndarray of psi values for each variable
        '''
        def psi(expected_array, actual_array, buckets):
            '''Calculate the PSI for a single variable
            Args:
            expected_array: numpy array of original values
            actual_array: numpy array of new values, same size as expected
            buckets: number of percentile ranges to bucket the values into
            Returns:
            psi_value: calculated PSI value
            '''
            def scale_range (input, min, max):
                input += -(np.min(input))
                input /= np.max(input) / (max - min)
                input += min
                return input
            breakpoints = np.arange(0, buckets + 1) / (buckets) * 100
            if buckettype == 'bins':
                breakpoints = scale_range(breakpoints, np.min(expected_array), np.max(expected_array))
            elif buckettype == 'quantiles':
                breakpoints = np.stack([np.percentile(expected_array, b) for b in breakpoints])
            expected_percents = np.histogram(expected_array, breakpoints)[0] / len(expected_array)
            actual_percents = np.histogram(actual_array, breakpoints)[0] / len(actual_array)
            def sub_psi(e_perc, a_perc):
                '''Calculate the actual PSI value from comparing the values.
                Update the actual value to a very small number if equal to zero
                '''
                if a_perc == 0:
                    a_perc = 0.0001
                if e_perc == 0:
                    e_perc = 0.0001
                value = (e_perc - a_perc) * np.log(e_perc / a_perc)
                return(value)
            psi_value = np.sum(sub_psi(expected_percents[i], actual_percents[i]) for i in range(0, len(expected_percents)))
            return(psi_value)

        if len(expected.shape) == 1:
            psi_values = np.empty(len(expected.shape))
        else:
            psi_values = np.empty(expected.shape[axis])

        for i in range(0, len(psi_values)):
            if len(psi_values) == 1:
                psi_values = psi(expected, actual, buckets)
            elif axis == 0:
                psi_values[i] = psi(expected[:,i], actual[:,i], buckets)
            elif axis == 1:
                psi_values[i] = psi(expected[i,:], actual[i,:], buckets)
        return(psi_values)
