from .ReadData import readData
from .EDAData import edaData
from .VarData import varData
from .ModelData import modelData
from .ResultsData import resultsData
from .ShapPlots import shapPlots
from .CalcPSI import calcPSI

import openpyxl as op
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN

from sklearn.metrics import roc_auc_score

import pandas as pd
import joblib
import os
import shutil
import site
from datetime import datetime
from tqdm import tqdm


def run_main(filename,target,vars_list=[],tot_cols=10,exclude_cols=[],exclude_suffix=[],exclude_prefix=[]):
    """
    Main function for the documentation

    Parameters
    ----------
    filename : str
            Input File Name
    target : str
            Target column
    vars_list : List of str
            List of variables to be selected
    tot_cols : int, default 10
            Total Columns to be selected from dataset
    exclude_cols : List of str, optional
                List of columns to be excluded in the dataframe
    exclude_suffix : List of str, optional
            List of columns with mentioned suffix to be excluded in the dataframe
    exclude_prefix : List of str, optional
            List of columns with mentioned prefix to be excluded in the dataframe

    Returns
    -------
    None
    """
    print("Process start time: ", datetime.now())

    ### Step 1. Reading the Data
    print("\nInput File Name:",filename)
    print("\nReading the input file started...")
    for i in tqdm (range (1), desc="Processing Step 1: Reading the Data..."):
        in_data = readData.readFile(filename) 
    print("\nReading the input file completed...")
    print("\nOutput File Name: documentation_final.xlsx")

    ## Step 2: Loading the template document
    print("\nLoading the template document started...")
    for i in tqdm (range (1), desc="Processing Step 2: Loading the template document..."):
        for pkg_dir in site.getsitepackages():
            try:
                base_path = pkg_dir+'/mladocs'
                path = base_path+'/documentation.xlsx.py'
                cwd = os.getcwd()
                out_path = cwd+'/documentation.xlsx'
                shutil.copyfile(path, out_path)
            except:
                print("\nFile not found in the directory:",pkg_dir)
    print("\nLoading the template document completed...")

    ### Step 3: Data statistics for Raw data
    print("\nWriting raw statistics data to excel started...")
    for i in tqdm (range (1), desc="Processing Step 3: Data statistics for Raw data..."):
        xfile = load_workbook('documentation.xlsx')    
        sheet = xfile['1. Statistics of Raw Data']    ## Access Raw Statistics sheet    
        sheet['B4']=filename    ## Fill in values
        sheet['B4'].font = Font(size=14,italic=True)
        sheet['B5']=target  ## Fill in values
        sheet['B5'].font = Font(size=14,italic=True)    
        xfile.save('documentation_final.xlsx') ## Save
        edaData().stats(in_data,data_level='raw') ## Summary stats
        edaData().tgtrate(in_data,target)   ## Events Rate
    print("\nWriting raw statistics data to excel completed...")

    ### Step 4: Features Selection and Null Handling
    print("\nSelecting Features and Hanlding Null/Unique values started...")
    for i in tqdm (range (1), desc="Processing Step 4: Features Selection and Null Handling..."):   
        keep_cols = edaData().excludecols(data=in_data,target=target,custom_rm_lst=exclude_cols,custom_rm_suffix=exclude_suffix,custom_rm_prefix=exclude_prefix)
        in_data = in_data[keep_cols]
        missing_value_df = edaData().nulldrop(data=in_data)
        print("\nMissing Values Dataframe:")
        print(missing_value_df)
        numeric, categorical = edaData().variableType(in_data, target=target)
        in_data = edaData().nullfill(data=in_data,target=target,num_cols=numeric,categorical_cols=categorical)
        in_data = edaData().uniquedrop(data=in_data)
    print("\nSelecting Features and Hanlding Null/Unique values completed...")

    ### Step 5. Categorical Encoding
    print("\nCategorical Encoding started...")
    for i in tqdm (range (1), desc="Processing Step 5. Categorical Encoding..."): 
        numeric, categorical = edaData().variableType(in_data, target=target)
        in_data = edaData().categoricalIndexer(in_data,categorical) ## Encoding Categorical Values
    print("\nCategorical Encoding completed...")

    ### Step 6. Variable Reduction   
    print("\nVariable Reduction started...")
    for i in tqdm (range (1), desc="Processing Step 6. Variable Reduction..."):
        if len(vars_list)==0:
            vars_list = varData().modelvariables(in_data,in_data.columns,target,exclude_cols,tot_cols,init_iv=True)
            print("Final variables List: ", vars_list)
    print("\nVariable Reduction completed...")

    ### Step 7: Data statistics for Final data
    print("\nWriting final statistics data to excel started...")
    for i in tqdm (range (1), desc="Processing Step 7: Data statistics for Final data..."):
        final_df = in_data[vars_list+[target]]
        edaData().stats(final_df,data_level='clean')
        xfile = load_workbook('documentation_final.xlsx')
        sheet = xfile['2. Statistics of Clean Data']
        thin_border = Border(
                            left=Side(border_style=BORDER_THIN, color='00000000'),
                            right=Side(border_style=BORDER_THIN, color='00000000'),
                            top=Side(border_style=BORDER_THIN, color='00000000'),
                            bottom=Side(border_style=BORDER_THIN, color='00000000'))
        for r in range(25,25+final_df.shape[1]):
            for c in range(2,15):
                sheet.cell(row=r,column=c).border = thin_border
        xfile.save('documentation_final.xlsx')
    print("\nWriting final statistics data to excel completed...")

    ### Step 8: Variable Binning and Bivariate Analysis on Final Variables
    print("\nVariable Binning and Bivariate Analysis on Final Variables started...")
    for i in tqdm (range (1), desc="Processing Step 8: Variable Binning and Bivariate Analysis on Final Variables..."):
        numeric, categorical = edaData().variableType(final_df, target=target)

        ## Variable Binning
        xfile = load_workbook('documentation_final.xlsx')
        sheet = xfile['3. Variable Binning']
        sheet['B4']=len(vars_list)
        sheet['B4'].font = Font(size=14,italic=True)
        sheet['B5']=len(numeric)
        sheet['B5'].font = Font(size=14,italic=True)
        sheet['B6']=str(numeric)
        sheet['B6'].font = Font(size=14,italic=True)
        sheet['B7']=len(categorical)
        sheet['B7'].font = Font(size=14,italic=True)
        sheet['B8']=str(categorical)
        sheet['B8'].font = Font(size=14,italic=True)
        xfile.save('documentation_final.xlsx')

        ## Bivariate Analysis
        xfile = load_workbook('documentation_final.xlsx')
        sheet = xfile['4. Bivariate Analysis']
        sheet['B2']=target
        sheet['B2'].font = Font(size=14,italic=True)
        sheet['B3']=len(vars_list)
        sheet['B3'].font = Font(size=14,italic=True)
        sheet['B4']=str(vars_list)
        sheet['B4'].font = Font(size=14,italic=True)
        xfile.save('documentation_final.xlsx')    

        mono_df = varData().modelvariables(final_df,final_df.columns,target,exclude_cols)
        print("\nVariable Binning Data:")
        print(mono_df)
        # Remove Bivariates directory
        shutil.rmtree('bivariates')
    print("\nVariable Binning and Bivariate Analysis on Final Variables completed...")

    ### Step 9. Model Building
    print("\nModel Building started...")
    for i in tqdm (range (1), desc="Processing Step 9. Model Building..."):
        X_train, X_test, y_train, y_test, model_results_df = modelData().modelbuild(data=final_df,cols=vars_list,target=target,split_size=0.70)
        print(model_results_df)
    print("\nModel Building completed...")

    ### Step 10. Model Evaluation
    print("\nModel Evaluation started...")
    for i in tqdm (range (1), desc="Processing Step 10. Model Evaluation..."):
        #Logistic Regression
        print("\nGenerating evaluation results for Logistic Regression Model...")
        lr_model = joblib.load('model_results/logreg_model.pkl')
        metrics_df = pd.DataFrame(columns=['Type of Data','Algorithm','Parameters','Accuracy Score','Precision Score','Recall Score','F1 Score','Confusion Matrix','Area under the ROC'])
        y_train_pred,metrics_df = modelData().modeleval(lr_model,X_train,y_train,metrics_df,data_type='train')
        rank_df = resultsData().genRank(y_train,y_train_pred)
        decile_table = resultsData().decilesData(data=rank_df[['tgt','Rank Order']],target="tgt", score="Rank Order")
        lr_spread = decile_table['SPREAD'].max()
        xfile = load_workbook('documentation_final.xlsx')
        sheet = xfile['10. Benchmark']
        for row, events, non_events in zip(range(8, 18), decile_table['TARGET'], decile_table['NONTARGET']):
            sheet['D'+str(row)] = non_events+events  ## Fill in Decile Values
            sheet['E'+str(row)] = events
        xfile.save('documentation_final.xlsx')

        #Random Forest
        print("\nGenerating evaluation results for Random Forest Model...")
        rf_model = joblib.load('model_results/rf_model.pkl')
        metrics_df = pd.DataFrame(columns=['Type of Data','Algorithm','Parameters','Accuracy Score','Precision Score','Recall Score','F1 Score','Confusion Matrix','Area under the ROC'])
        y_train_pred,metrics_df = modelData().modeleval(rf_model,X_train,y_train,metrics_df,data_type='train')
        rank_df = resultsData().genRank(y_train,y_train_pred)
        decile_table = resultsData().decilesData(data=rank_df[['tgt','Rank Order']],target="tgt", score="Rank Order")
        rf_spread = decile_table['SPREAD'].max()
        xfile = load_workbook('documentation_final.xlsx')
        sheet = xfile['10. Benchmark']        
        for row, events, non_events in zip(range(26, 36), decile_table['TARGET'], decile_table['NONTARGET']):
            sheet['D'+str(row)] = non_events+events ## Fill in Decile Values
            sheet['E'+str(row)] = events
        xfile.save('documentation_final.xlsx')

        #XGB
        print("\nGenerating evaluation results for XGB Model...")
        xgb_model = joblib.load('model_results/xgb_model.pkl')      
        metrics_df = pd.DataFrame(columns=['Type of Data','Algorithm','Parameters','Accuracy Score','Precision Score','Recall Score','F1 Score','Confusion Matrix','Area under the ROC'])
        y_train_pred,metrics_df = modelData().modeleval(xgb_model,X_train,y_train,metrics_df,data_type='train')
        y_test_pred,metrics_df = modelData().modeleval(xgb_model,X_test,y_test,metrics_df,data_type='test')
        rank_df = resultsData().genRank(y_train,y_train_pred)
        decile_table = resultsData().decilesData(data=rank_df[['tgt','Rank Order']],target="tgt", score="Rank Order")
        xgb_spread = decile_table['SPREAD'].max()
        xfile = op.load_workbook('documentation_final.xlsx')
        sheet = xfile['10. Benchmark']    
        for row, events, non_events in zip(range(44, 54), decile_table['TARGET'], decile_table['NONTARGET']):
            sheet['D'+str(row)] = non_events+events ## Fill in Decile Values
            sheet['E'+str(row)] = events
        xfile.save('documentation_final.xlsx')
    print("\nModel Evaluation completed...")

    if (lr_spread > rf_spread) and (lr_spread > xgb_spread):
        model = lr_model
        final_model = 'logit'
    elif (rf_spread > lr_spread) and (lr_spread > xgb_spread):
        model = rf_model
        final_model = 'rf'
    else:
        model = xgb_model
        final_model = 'xgb'
    
    ### Step 11. Data Deciling
    print("\nData Deciling started...")
    for i in tqdm (range (1), desc="Processing Step 11. Data Deciling..."):
        vif_df = varData().vifvalues(final_df[vars_list],model,select_model=final_model)   ## Variables Importance - VIF
        print(vif_df)        
        xfile = load_workbook('documentation_final.xlsx')
        sheet = xfile['7. Decile Tables']
        for row, events, non_events in zip(range(5, 15), decile_table['TARGET'], decile_table['NONTARGET']):
            sheet['D'+str(row)] = non_events+events
            sheet['E'+str(row)] = events
        xfile.save('documentation_final.xlsx')
        rank_df = resultsData().genRank(y_test,y_test_pred)
        decile_table = resultsData().decilesData(data=rank_df[['tgt','Rank Order']],target="tgt", score="Rank Order")
        xfile = load_workbook('documentation_final.xlsx')
        sheet = xfile['7. Decile Tables']    
        for row, events, non_events in zip(range(22, 32), decile_table['TARGET'], decile_table['NONTARGET']):
            sheet['D'+str(row)] = non_events+events ## Fill in Decile Values
            sheet['E'+str(row)] = events
        xfile.save('documentation_final.xlsx')
    print("\nData Deciling completed...")

    ### Step 12. Gains and Lift
    print("\nGenerating Gains and Lift started...")
    for i in tqdm (range (1), desc="Processing Step 12. Gains and Lift..."):
        resultsData().gains(data=rank_df[['tgt','score','Rank Order']],target="tgt",score="score")
        # Remove Gains/Lift directory
        shutil.rmtree('gains_lifts')
    print("\nGenerating Gains and Lift completed...")

    ### Step 13. PSI Check
    print("\nPSI Check started...")
    for i in tqdm (range (1), desc="Processing Step 13. PSI Check..."):
        psi_values = resultsData().calculate_psi(rank_df['tgt'].dropna(), rank_df['score'].dropna())
        proba = model.predict_proba(X_train)
        proba_test = model.predict_proba(X_test)        
        psi = calcPSI(y_train, proba[:,1])  ## Initialize PSI class        
        psi.psi_dev()   ## Calculate bins from train        
        final_psi = psi.psi_test(y_test, proba_test[:,1])   ## Calculate bin counts for test
        print(final_psi)
        xfile = load_workbook('documentation_final.xlsx')
        sheet = xfile['9. Model Stability']
        ## Populate values for PSI calculation
        for row, max_prob, min_prob, dev_count, test_count in zip(range(25, 35), final_psi['MAX_SCORE'], final_psi['MIN_SCORE'], final_psi['TOTAL'], final_psi['TEST_COUNTS']):
            sheet['B'+str(row)] = max_prob
            sheet['C'+str(row)] = min_prob
            sheet['D'+str(row)] = dev_count
            sheet['F'+str(row)] = test_count
        xfile.save('documentation_final.xlsx')
        print("PSI Values: ",psi_values)
    print("\nPSI Check completed...")    

    ### Step 14. Shap Plots
    print("\nShap Plots Generation started...")
    for i in tqdm (range (1), desc="Processing Step 14. Shap Plots..."):
        train_sample = X_train.sample(n=100)
        shap_plots = shapPlots(model, train_sample)
        shap_plots.summaryPlot(save_plot="shap_plots")
        shap_plots.abs_shap(save_plot="shap_plots")
        print("\nInserting Shap Plots to excel...")
        xfile = load_workbook('documentation_final.xlsx')
        save_plot="shap_plots"
        sheet = xfile['11. Shapley Values']
        img = op.drawing.image.Image(save_plot+"/summary_plot.png")
        img.anchor = 'D10'
        img.height = 1200
        img.width= 1500
        sheet.add_image(img)
        img = op.drawing.image.Image(save_plot+"/abs_shap.png")
        img.anchor = 'D80'
        img.height = 1200
        img.width= 1500
        sheet.add_image(img)
        xfile.save('documentation_final.xlsx')
        # Remove Shap Plots Directory
        shutil.rmtree('shap_plots')
    print("\nShap Plots Generation completed...")

    ### Step 15. Other Observations
    print("\nWriting the other observations to excel started...")
    for i in tqdm (range (1), desc="Processing Step 15. Other Observations..."):
        xfile = load_workbook('documentation_final.xlsx')
        sheet = xfile['12. Other Observations']

        vif_high_list = list(vif_df[vif_df['VIF Scores']>=5]['Variables'].values)
        if len(vif_high_list) != 0:
            sheet['A5'] = "There are variables who have VIF scores higher than 5. Please remove these variables for better results."
            sheet['A6'] = "The following variables have VIF values greater than 5. Please check the '5. Variable Importance' sheet for the VIF values and their importance:"
            sheet['A7'] = str(vif_high_list)
        else:
            sheet['A5'] = "There are no variables who have VIF scores higher than 5."

        imp_high_list = list(vif_df[vif_df['Importance']>=0.4]['Variables'].values)
        if len(imp_high_list) != 0:
            sheet['A15'] = "There are variables who have Importance higher than 0.4. The model will skew towards these variables."
            sheet['A16'] = "The following variables have Importance greater than 0.4. Please check the '5. Variable Importance' sheet for the VIF values and their importance:"
            sheet['A17'] = str(imp_high_list)
        else:
            sheet['A15'] = "There are no variables who have Importance higher than 0.4"

        if roc_auc_score(y_train, y_train_pred) > 0.95:
            sheet['A25'] = "The ROC value for train data is " + str(roc_auc_score(y_train, y_train_pred)) + " which is high. Please reverify the variables."
        elif roc_auc_score(y_test, y_test_pred) > 0.95:
            sheet['A25'] = "The ROC value for test data is " + str(roc_auc_score(y_test, y_test_pred)) + " which is high. Please reverify the variables."
        else:
            sheet['A25'] = "The ROC value looks good for both train and test data at " + str(roc_auc_score(y_train, y_train_pred)) + " and " + str(roc_auc_score(y_test, y_test_pred)) + " respectively."

        psi_values
        if psi_values < 0.1:
            sheet['A35'] = "The PSI value is " + str(psi_values) + " which is good."
        elif 0.1 <= psi_values <= 0.25:
            sheet['A35'] = "The PSI value is " + str(psi_values) + " which is high. Please investigate the current samples."
        else:
            sheet['A35'] = "The PSI value is " + str(psi_values) + " which is very high.  It is strongly recommended to build a new model using more recent sample."

        xfile.save('documentation_final.xlsx')
    print("\nWriting the other observations to excel completed...")

    print("\nDocumentation Generation Completed...")

    print("\nProcess end time: ", datetime.now())

    return None
 
# run_main(filename='input_file.csv',target='rgu_xmobile',vars_list=[],tot_cols=10,exclude_cols=['CUSTOMER_ACCOUNT_ID', 'ACCOUNT_NUMBER', 'ACCOUNTID', 'ACCOUNT', 'HOUSEKEY', 'BILLER_EMAIL'],exclude_suffix=['_date', '_id'],exclude_prefix=[])