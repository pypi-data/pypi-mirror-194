"""
Variable Importance Data Module
"""

import pandas as pd
import numpy as np
import openpyxl as op
from openpyxl import load_workbook
import os
from openpyxl.styles import Font,PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_THIN
import seaborn as sns
import matplotlib.pyplot as plt
from varclushi import VarClusHi
import scipy.stats as stats
from statsmodels.stats.outliers_influence import variance_inflation_factor
from tqdm import tqdm

class varData:
    """Class to identify, reduce, visualize the variables data of the dataframe"""

    def varReduce(self,data,max_eigval=1,max_cluster=20):
        """
        Reduce the dimensionality of the independent variables using Variable Clustering, excluding the target variable.

        Parameters
        ----------
        data : dataframe
                Input dataframe.
        max_eigval : int, optional, default value is 1
                Maximum Eigenvalue to consider
        max_cluster : int, optional, default value is 20
                Maximum number of variable clusters

        Returns
        -------
        vars_list : List of str
                    Returns the final set of variables from Variable Clustering.
        """
        print("\nVariable Reduction using Variable Clustering started...")
        new_data_vc = VarClusHi(data,maxeigval2=max_eigval,maxclus=max_cluster)
        new_data_vc.varclus()

        print("\nClusters Info Table...")
        print(new_data_vc.info)

        print("\nSelecting the final set of variables from the R-Square ratio Table...")
        print(new_data_vc.rsquare)
        vars_list = list(new_data_vc.rsquare.loc[new_data_vc.rsquare.groupby('Cluster')['RS_Ratio'].idxmin()]['Variable'])

        print("\nFinal set of variables selected are: ")
        print(vars_list)

        return vars_list

    def monobin(self,Y, X, n = 10):
        """
        Monotonic Binning of numerical variables

        Parameters
        ----------
        Y : Series
            Target column.
        X : Series
            Independent column.
        n : int, optional, default value is 10
                Number of bins

        Returns
        -------
        mono_df : dataframe
                Returns the numerical variable binning dataframe
        """
        print("\nGenerating the dataframe with the input and target...")
        df1 = pd.DataFrame({"X": X, "Y": Y})
        notmiss = df1.fillna(0)
        r = 0
        print("\nMonotonic Binning starts...")
        while np.abs(r) < 1:
            try:
                if n==0:
                    d1 = pd.DataFrame({"X": X, "Y": Y, "Bucket": pd.qcut(X, n,duplicates='drop')})
                else:
                    d1 = pd.DataFrame({"X": X, "Y": Y, "Bucket": pd.qcut(X, n)})
                d2 = d1.groupby('Bucket', as_index=True)
                r, p = stats.spearmanr(d2.mean().X, d2.mean().Y)
                n = n - 1
            except Exception as e:
                n =  n - 1

        if len(d2) == 1:
            print("\nBinning for columns with few unique values...")
            n = 5
            bins = np.quantile(notmiss.X, np.linspace(0, 1, n))
            bins = np.insert(bins, 0, 0.5)
            bins[1] = bins[1]-(bins[1]/2)
            d1 = pd.DataFrame({"X": notmiss.X, "Y": notmiss.Y, "Bucket": pd.cut(notmiss.X, np.unique(bins),include_lowest=True)})
            d2 = d1.groupby('Bucket', as_index=True)

        print("\nGenerating the binning dataframe")
        mono_df = pd.DataFrame({},index=[])
        mono_df["MIN_VALUE"] = d2.min().X
        mono_df["MAX_VALUE"] = d2.max().X
        mono_df["COUNT"] = d2.count().Y
        mono_df["EVENT"] = d2.sum().Y
        mono_df["NONEVENT"] = d2.count().Y - d2.sum().Y
        mono_df=mono_df.reset_index(drop=True)
        mono_df["EVENT_RATE"] = mono_df.EVENT/mono_df.COUNT
        mono_df["NON_EVENT_RATE"] = mono_df.NONEVENT/mono_df.COUNT
        mono_df["WOE"] = np.log(mono_df.EVENT_RATE/mono_df.NON_EVENT_RATE)
        mono_df["IV"] = (mono_df.EVENT_RATE-mono_df.NON_EVENT_RATE)*np.log(mono_df.EVENT_RATE/mono_df.NON_EVENT_RATE)
        mono_df["VAR_NAME"] = "VAR"
        mono_df = mono_df[['VAR_NAME','MIN_VALUE', 'MAX_VALUE', 'COUNT', 'EVENT', 'EVENT_RATE', 'NONEVENT', 'NON_EVENT_RATE', 'WOE', 'IV']]
        mono_df = mono_df.replace([np.inf, -np.inf], 0)
        mono_df.IV = mono_df.IV.sum()


        print("\nMonotonic Binning ends...")
        return(mono_df)

    def charbin(self,Y, X):
        """
        Character Binning of categorical variables

        Parameters
        ----------
        Y : Series
            Target column.
        X : Series
            Independent column.

        Returns
        -------
        char_df : dataframe
                Returns the numerical variable binning dataframe
        """
        print("\nGenerating the dataframe with the input and target...")
        df1 = pd.DataFrame({"X": X, "Y": Y})
        df2 = df1.groupby('X',as_index=True)
        var = X.name

        print("\nCharacter Binning starts...")
        char_df = pd.DataFrame({},index=[])
        char_df["COUNT"] = df2.count().Y
        char_df["MIN_VALUE"] = df2.apply(lambda x: x.name)
        char_df["MAX_VALUE"] = df2.apply(lambda x: x.name)
        char_df["EVENT"] = df2.sum().Y
        char_df["NONEVENT"] = df2.count().Y - df2.sum().Y
        char_df["EVENT_RATE"] = char_df.EVENT/char_df.COUNT
        char_df["NON_EVENT_RATE"] = char_df.NONEVENT/char_df.COUNT
        char_df["WOE"] = np.log(char_df.EVENT_RATE/char_df.NON_EVENT_RATE)
        char_df["IV"] = (char_df.EVENT_RATE-char_df.NON_EVENT_RATE)*np.log(char_df.EVENT_RATE/char_df.NON_EVENT_RATE)
        char_df["VAR_NAME"] = var
        char_df = char_df[['VAR_NAME','MIN_VALUE', 'MAX_VALUE', 'COUNT', 'EVENT', 'EVENT_RATE', 'NONEVENT', 'NON_EVENT_RATE', 'WOE', 'IV']]
        for i in char_df.columns:
            if char_df[i].dtypes != 'object':
                char_df[i] = char_df[i].replace([np.inf, -np.inf], 0)
        char_df.IV = char_df.IV.sum()
        char_df = char_df.reset_index(drop=True)

        print("\nCharacter Binning ends...")

        return(char_df)

    def datavars(self,df1, target, exclusion_cols):
        """
        Generating the Variable Binning dataframe with WOE and IV values

        Parameters
        ----------
        df1 : dataframe
                Input dataframe.
        target : Series
                Target column
        exclusion_cols : List of str
                List of columns to be excluded from binning

        Returns
        -------
        iv_df : dataframe
                    Returns the dataframe containing the variables binning details.
        """
        print("\nBinning both numerical and categorical variables started...")
        x = df1.dtypes.index
        count = -1
        df2 = df1.select_dtypes(include=[np.number])
        print(df2)

        for i in x:
            if i.upper() not in exclusion_cols:
                if i in df2 and len(np.unique(df1[i])) > 10:
                    ### Numerical Variables Binning
                    print("\nBinning for the column: ",i)
                    conv = self.monobin(target, df1[i])
                    conv["VAR_NAME"] = i
                    count = count + 1
                else:
                    ### Categorical Variables Binning
                    print("\nBinning for the column: ",i)
                    conv = self.charbin(target, df1[i])
                    conv["VAR_NAME"] = i
                    conv = conv.sort_values('EVENT_RATE')
                    count = count + 1
                ### Generating the bins
                if count == 0:
                    iv_df = conv
                else:
                    # iv_df = iv_df.append(conv,ignore_index=True)
                    iv_df = pd.concat([iv_df,conv])
        print("\nBinning both numerical and categorical variables completed...")
        return(iv_df)

    def modelvariables(self,data,inputs,target,cols,tot_cols=0,init_iv=False):
        """
        Perform Variables Binning and Bivariate Analysis of each independent variable with the target.

        Parameters
        ----------
        data : dataframe
                Input dataframe.
        inputs : List of str
                List of dataframe columns
        target : str
                Target column name
        cols : List of str
                List of columns to be excluded from binning

        Returns
        -------
        final_iv : dataframe
                    Returns the dataframe containing the variables binning details.
        """
        print("\nVariable Binning Started...")
        data = data[inputs]
        final_iv_df = self.datavars(data,data[target],cols)
        final_iv = final_iv_df[final_iv_df.VAR_NAME != target].copy()
        grouped = final_iv.groupby(['VAR_NAME'])
        i = 1
        print(final_iv)

        iv_df = final_iv[['VAR_NAME','IV']].copy()
        iv_df.drop_duplicates(inplace=True)
        iv_df.sort_values(by='IV', ascending=False,inplace=True)
        
        if tot_cols == 0:
            if len(iv_df['IV']) >= 10:
                cols_list = list(iv_df['VAR_NAME'].head(10))
            elif (len(iv_df['IV']) <= 10) and (len(iv_df['IV']) >= 5):
                cols_list = list(iv_df['VAR_NAME'].head(5))
        else:
            cols_list = list(iv_df['VAR_NAME'].head(tot_cols))

        if init_iv == True:
            print('Exiting binning...')   
            return(cols_list)

        print("\nWriting Information Value to excel...")
        with pd.ExcelWriter("documentation_final.xlsx",mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
            iv_df.to_excel(writer, sheet_name='3. Variable Binning',startrow=13,startcol=13,index=False,header=False)

        print("\nBivariate Analysis Started...")
        for key, group in grouped:
            plt.figure(figsize = (8, 10))
            group['MIN_VALUE']=group['MIN_VALUE'].apply(lambda x: round(x,2))
            ax = sns.barplot(group,x='MIN_VALUE',y='EVENT_RATE')
            ax.set_title(str(key) + " vs " + str(target))
            ax.set_xlabel(key)
            ax.set_ylabel(str(target) + " " + " %")
            rects = ax.patches
            for rect in rects:
                height = rect.get_height()
                ax.text(rect.get_x()+rect.get_width()/2., 1.01*height, str(round(height*100,1)) + '%', ha='center', va='bottom', fontweight='bold')
            newpath = r'bivariates'
            if not os.path.exists(newpath):
                os.makedirs(newpath)
                plt.savefig(newpath+"/"+str(i)+'.png')
            else:
                plt.savefig(newpath+"/"+str(i)+'.png')
            i = i+1

        print("\nVariable Binning Dataframe Generated...")

        print("\nWriting Variable Binning data to excel starts...")
        with pd.ExcelWriter("documentation_final.xlsx",mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
            final_iv.to_excel(writer, sheet_name='3. Variable Binning',startrow=13,startcol=0,index=False)

        xfile = load_workbook('documentation_final.xlsx')
        sheet = xfile['3. Variable Binning']
        for c in range(1,11):
                        sheet.cell(row=14, column=c).font = Font(size=14,bold=True)
                        sheet.cell(row=14, column=c).fill = PatternFill(start_color="538DD5", end_color="538DD5", fill_type = "solid")
        thin_border = Border(
                        left=Side(border_style=BORDER_THIN, color='00000000'),
                        right=Side(border_style=BORDER_THIN, color='00000000'),
                        top=Side(border_style=BORDER_THIN, color='00000000'),
                        bottom=Side(border_style=BORDER_THIN, color='00000000'))
        for r in range(14,15+final_iv.shape[0]):
            for c in range(1,11):
                sheet.cell(row=r,column=c).border = thin_border
        for r in range(14,14+iv_df.shape[0]):
            for c in range(14,16):
                sheet.cell(row=r,column=c).border = thin_border
        xfile.save('documentation_final.xlsx')
        print("\nWriting Variable Binning data to excel ends...")

        print("\nBivariate Analysis in excel starts...")
        xfile = load_workbook('documentation_final.xlsx')
        sheet = xfile['4. Bivariate Analysis']
        i = 1
        j = 6
        for key, group in grouped:
            img = op.drawing.image.Image(newpath+"/"+str(i)+'.png')
            img.anchor = 'C'+str(j)
            sheet.add_image(img)
            i=i+1
            j=j+60
        xfile.save('documentation_final.xlsx')

        
        print("\nBivariate Analysis in excel ends...")

        return (final_iv)

    def vifvalues(self,data,model,select_model):
        """
        Calculate Variance Inflation Factor

        Parameters
        ----------
        data : dataframe
                Input dataframe.
        model : Class Object
                Model Object
        select_model : str
                Type of Model

        Returns
        -------
        vif_scores : dataframe
                    Returns the dataframe containing the Variance Inflation Factor.
        """
        print("\nVariance Inflation Factor calculation starts...")

        vif_scores = pd.DataFrame()
        vif_scores["Variables"] = data.columns
        if select_model == 'logit':
            vif_scores["Importance"] = model.coef_[0]
        else:
            vif_scores["Importance"] = model.feature_importances_

        print("\nCalculating VIF for each feature")
        vif_scores["VIF Scores"] = [variance_inflation_factor(data.values, i) for i in range(len(data.columns))]

        print("\nVariance Inflation Factor calculated...")

        print("\nWriting Variance Inflation Factor to excel starts...")
        vif_scores = vif_scores[['Variables','Importance','VIF Scores']].sort_values('Importance', ascending=False)
        with pd.ExcelWriter("documentation_final.xlsx",mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
            vif_scores.to_excel(writer, sheet_name='5. Variable Importance',startrow=2,startcol=0,header=False)

        xfile = load_workbook('documentation_final.xlsx')
        sheet = xfile['5. Variable Importance']
        thin_border = Border(
                                left=Side(border_style=BORDER_THIN, color='00000000'),
                                right=Side(border_style=BORDER_THIN, color='00000000'),
                                top=Side(border_style=BORDER_THIN, color='00000000'),
                                bottom=Side(border_style=BORDER_THIN, color='00000000'))
        for r in range(3,len(data.columns)+3):
                for c in range(1,5):
                        sheet.cell(row=r, column=c).border = thin_border
        xfile.save('documentation_final.xlsx')
        print("\nWriting Variance Inflation Factor to excel ends...")

        return (vif_scores)
