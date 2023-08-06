"""
EDA Data Module
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_THIN
import category_encoders as ce

class edaData:
    """Class to perform EDA on the given data"""

    def tgtrate(self,data,target):
        """
        Calculate the event and non-event rate of the target.

        Parameters
        ----------
        data : dataframe
                Input dataframe.
        target : str
                Target column name.

        Returns
        -------
        None
        """
        print("\nTarget Column: ",target)
        event_rate = (data[target][data[target]==1].count() / len(data))*100
        nonevent_rate = (data[target][data[target]==0].count() / len(data))*100

        print("\nRate of Non-Events:",nonevent_rate)
        print("\nRate of Events:",event_rate)

        print("\nWriting events and non-events rate to excel...")
        xfile = load_workbook('documentation_final.xlsx')
        sheet = xfile['1. Statistics of Raw Data']
        sheet['B8']=str(event_rate)
        sheet['B8'].font = Font(size=14,italic=True)
        sheet['B9']=str(nonevent_rate)
        sheet['B9'].font = Font(size=14,italic=True)
        xfile.save('documentation_final.xlsx')

        return None

    def stats(self,data,data_level):
        """
        View the dataframe statistics.

        Parameters
        ----------
        data : dataframe
                Input dataframe.        
        data_level : str
                Level of data as 'raw' or 'clean'

        Returns
        -------
        None
        """
        print("\nTop 5 rows of the dataframe")
        print(data.head())

        print("\nShape of the dataframe")
        print(data.shape)

        print("\nInfo on the dataframe")
        print(data.info())

        print("\nSummary Statistics of the dataframe")
        print(data.describe(include="all"))
        dt = data.describe(include="all").T
        dt['datatype'] = data.dtypes
        dt['missing_values'] = data.isnull().sum() * 100 / len(data)

        if data_level == 'raw':           
                with pd.ExcelWriter("documentation_final.xlsx",mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
                        print("\nWriting header rows to excel...")
                        data.head().to_excel(writer, sheet_name="1. Statistics of Raw Data", startrow=13,startcol=0,index=False)
                        print("\nWriting summary rows to excel...")
                        dt.to_excel(writer, sheet_name='1. Statistics of Raw Data',startrow=28,startcol=0,header=False)                
                print("\nFormatting Shapes display...")
                xfile = load_workbook('documentation_final.xlsx')
                sheet = xfile['1. Statistics of Raw Data']
                sheet['B6'] = data.shape[0]
                sheet['B6'].font = Font(size=14,italic=True)
                sheet['B7'] = data.shape[1]
                sheet['B7'].font = Font(size=14,italic=True)

                print("\nFormatting headers display...")
                for c in range(1,data.shape[1]+1):
                        sheet.cell(row=14, column=c).font = Font(size=14,bold=True)
                        sheet.cell(row=14, column=c).fill = PatternFill(start_color="538DD5", end_color="538DD5", fill_type = "solid")

                print("\nFormatting borders display...")
                thin_border = Border(
                                left=Side(border_style=BORDER_THIN, color='00000000'),
                                right=Side(border_style=BORDER_THIN, color='00000000'),
                                top=Side(border_style=BORDER_THIN, color='00000000'),
                                bottom=Side(border_style=BORDER_THIN, color='00000000'))
                for r in range(14,20):
                        for c in range(1,data.shape[1]+1):
                                sheet.cell(row=r, column=c).border = thin_border
                for r in range(29,data.shape[1]+1):
                        for c in range(2,15):
                                sheet.cell(row=r, column=c).border = thin_border

                xfile.save('documentation_final.xlsx')
        else:
                if 'count' not in dt.columns:
                        dt['count']=''
                if 'unique' not in dt.columns:
                        dt['unique']=''
                if 'top' not in dt.columns:
                        dt['top']=''
                if 'freq' not in dt.columns:
                        dt['freq']=''
                if 'mean' not in dt.columns:
                        dt['mean']=''
                if 'std' not in dt.columns:
                        dt['std']=''
                if 'min' not in dt.columns:
                        dt['min']=''
                if '25%' not in dt.columns:
                        dt['25%']=''
                if '50%' not in dt.columns:
                        dt['50%']=''
                if '75%' not in dt.columns:
                        dt['75%']=''
                if 'max' not in dt.columns:
                        dt['max']=''
                dt = dt[['count','unique','top','freq','mean','std','min','25%','50%','75%','max','datatype','missing_values']]                
                with pd.ExcelWriter("documentation_final.xlsx",mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
                        print("\nWriting header rows to excel...")
                        data.head().to_excel(writer, sheet_name='2. Statistics of Clean Data',startrow=9,startcol=0,index=False)
                        print("\nWriting summary rows to excel...")
                        dt.to_excel(writer, sheet_name='2. Statistics of Clean Data',startrow=24,startcol=0,header=False)

                print("\nFormatting Shapes display...")
                xfile = load_workbook('documentation_final.xlsx')
                sheet = xfile['2. Statistics of Clean Data']
                sheet['B4'] = data.shape[0]
                sheet['B4'].font = Font(size=14,italic=True)
                sheet['B5'] = data.shape[1]
                sheet['B5'].font = Font(size=14,italic=True)

                print("\nFormatting headers display...")
                for c in range(1,data.shape[1]+1):
                        sheet.cell(row=10, column=c).font = Font(size=14,bold=True)
                        sheet.cell(row=10, column=c).fill = PatternFill(start_color="538DD5", end_color="538DD5", fill_type = "solid")

                print("\nFormatting borders display...")
                thin_border = Border(
                                left=Side(border_style=BORDER_THIN, color='00000000'),
                                right=Side(border_style=BORDER_THIN, color='00000000'),
                                top=Side(border_style=BORDER_THIN, color='00000000'),
                                bottom=Side(border_style=BORDER_THIN, color='00000000'))
                for r in range(11,16):
                        for c in range(1,data.shape[1]+1):
                                sheet.cell(row=r, column=c).border = thin_border
                for r in range(25,data.shape[1]+1):
                        for c in range(2,15):
                                sheet.cell(row=r,column=c).border = thin_border

                xfile.save('documentation_final.xlsx')

        return None

    def variableType(self,data,target):
        """
        Determine numeric and categorical columns except target column

        Parameters
        ----------
        data : dataframe
                Input dataframe.
        target : str
                Target column name

        Returns
        -------
        numeric : List of str
                    Returns the list of numerical columns in the dataframe
        categorical : List of str
                    Returns the list of categorical columns in the dataframe
        """
        print("\nCollecting the columns for type identification...")
        variables = data.dtypes
        categorical = []
        numeric=[]

        for idx in range(0, len(variables)):
            if variables.index[idx] not in [target]:
                if variables[idx] == "object":
                    categorical.append(variables.index[idx])
                elif variables[idx] not in ['timestamp', 'date']:
                    numeric.append(variables.index[idx])
        print("\nCollected the numeric and categorical columns...")

        return numeric, categorical

    def cardinalityCount(self,data, columns, threshold):
        """
        Get distinct value counts for categorical variables

        Parameters
        ----------
        data : dataframe
                Input dataframe.
        columns : List of str
                List of categorical columns
        threshold : int
                Threshold value for the allowed cardinality in categorical columns

        Returns
        -------
        high_cardinal_vars : List of str
                    Returns the list of high cardinal columns
        zero_var_cardinal_vars : List of str
                    Returns the list of zero cardinal columns
        """
        ### Get counts
        print("\nCollecting the unique values in all columns...")
        result = data[columns].nunique().T
        result = result.reset_index().rename(columns={'index':'Variable', 0:'Count'})

        ### Get columns
        print("\nHigh Cardinal and Zero Cardinal columns...")
        high_cardinal_vars = list(result.loc[result['Count'] > threshold, 'Variable'])
        zero_var_cardinal_vars = list(result.loc[result['Count'] == 1, 'Variable'])

        return high_cardinal_vars, zero_var_cardinal_vars

    def excludecols(self,data,target,cardinality_threshold=120,custom_rm_lst=[],custom_rm_suffix=[],custom_rm_prefix=[]):
        """
        Remove the insignificant columns of the dataframe

        Parameters
        ----------
        data : dataframe
                Input dataframe.
        target : str
                Target column name
        cardinality_threshold : int, default value is 120
                Threshold value for the allowed cardinality in categorical columns
        custom_rm_lst : List of str, optional
                List of columns to be excluded in the dataframe
        custom_rm_suffix : List of str, optional
                List of columns with mentioned suffix to be excluded in the dataframe
        custom_rm_prefix : List of str, optional
                List of columns with mentioned prefix to be excluded in the dataframe

        Returns
        -------
        keep_cols : List of str
                    Returns the list of final columns in the dataframe
        """

        ### Select the required columns
        print("\nSelecting the required columns from the dataframe...")
        keep_cols = data.columns

        print(f"\nTotal number of variables selected are: {len(keep_cols)}")

        ### Excluding custom columns list
        print("\nExcluding the custom list of columns...")
        keep_cols = set(keep_cols) - set(custom_rm_lst)

        ### Removing columns with the suffixes mentioned in custom_rm_suffix
        print("\nRemoving custom list of columns with suffix...")
        custom_rm_suffix_lst = [var for suffix in custom_rm_suffix for var in keep_cols if var.endswith(suffix)]
        keep_cols = keep_cols - set(custom_rm_suffix_lst)

        ### Removing columns with the prefixes mentioned in custom_rm_prefix
        print("\nRemoving custom list of columns with prefix...")
        custom_rm_prefix_lst = [var for prefix in custom_rm_prefix for var in keep_cols if var.endswith(prefix)]
        keep_cols = keep_cols - set(custom_rm_prefix_lst)

        ### Removing columns with datatype date/timestamp
        print("\nRemoving columns with date/timestamp datatype...")
        all_vars = data.dtypes
        date_timestamp_rm = []
        for i in range(0, len(all_vars)):
            if all_vars[i] in ['timestamp', 'date']:
                date_timestamp_rm.append(all_vars[i][0])
        keep_cols = keep_cols - set(date_timestamp_rm)

        ### Collecting the Numerical and Categorical columns
        print("\nCollecting the Numerical and Categorical columns...")
        numeric, categorical = self.variableType(data, target)

        ### Removing variables with high cardinality
        print("\nRemoving variables with high cardinality...")
        high_cardinal_vars, zero_var_cardinal_vars = self.cardinalityCount(data, categorical, cardinality_threshold)
        keep_cols = keep_cols - set(high_cardinal_vars) - set(zero_var_cardinal_vars)

        print(f"\nFinal count of variables to be included are {len(keep_cols)}")

        return keep_cols

    def nulldrop(self,data, null_threshold=100.0):
        """
        Check and drop null values for mentioned threshold

        Parameters
        ----------
        data : dataframe
                Input dataframe.
        null_threshold : float, optional, default value is 100.0
                Count of Null values in the columns matching this threshold to be dropped

        Returns
        -------
        missing_value_df : dataframe
                Returns the dataframe with the columns and their corresponding nulls
        """
        print("\nCalculating the null value in each column...")
        percent_missing = data.isnull().sum() * 100 / len(data)

        print("\nCreating dataframe with null value details for each column...")
        missing_value_df = pd.DataFrame({'column_name': data.columns,'percent_missing': percent_missing})
        missing_value_df.sort_values('percent_missing', inplace=True)

        print("\nDropping the columns with nulls from threshold...")
        null_cols = list(missing_value_df[missing_value_df['percent_missing']>=null_threshold]['column_name'])
        data.drop(columns = null_cols,axis=1,inplace=True)     # Dropping Columns with the given null threshold

        percent_missing = data.isnull().sum() * 100 / len(data)
        missing_value_df = pd.DataFrame({'column_name': data.columns,'percent_missing': percent_missing})
        missing_value_df.sort_values('percent_missing', inplace=True)

        print("\nDataframe cleared of null values for the required columns...")
        return missing_value_df

    def nullfill(self,data,target,num_cols=[],categorical_cols=[]):
        """
        Fill null values in the columns

        Parameters
        ----------
        data : dataframe
                Input dataframe.
        target : str
                Target column name
        num_cols : List of str, optional
                List of custom numerical columns
        categorical_cols : List of str, optional
                List of custom categorical columns

        Returns
        -------
        data : dataframe
                Returns the dataframe without null values
        """

        ### Fill null values according to type of variable
        print("\nFilling the null values for numerical columns...")
        data[num_cols] = data[num_cols].fillna(value=0.0)

        print("\nFilling the null values for categorical columns...")
        data[categorical_cols] = data[categorical_cols].fillna(value='Other')

        return data

    def uniquedrop(self,data):
        """
        Check and drop columns with high unique values

        Parameters
        ----------
        data : dataframe
                Input dataframe.

        Returns
        -------
        data : dataframe
                Returns the dataframe with the high unique value columns removed
        """
        high_unique_cols = data.columns[data.nunique() <= 1]
        print("\nCount of columns with high unique values: ",len(high_unique_cols))

        data = data.loc[:, ~data.columns.isin(list(high_unique_cols))]

        print("\nDropped high unique value columns...")
        return data

    def categoricalIndexer(self, data, columns):
        """
        Encoding the categorical columns with binary integer

        Parameters
        ----------
        data : dataframe
                Input dataframe.
        columns : List of str
                List of categoical columns

        Returns
        -------
        encoded_data : dataframe
                Returns the dataframe with the categorical columns encoded
        """
        encoder = ce.BinaryEncoder(cols=columns)
        encoded_data = encoder.fit_transform(data)
        return encoded_data
