# import base64
import ast
import datetime
import json
import math
import os
import cx_Oracle
import mysql.connector
from urllib.parse import urlencode

import openpyxl
import requests
from django.conf import settings
from django.db.models import Max, Min
# import uuid
from rest_framework_tracking.models import APIRequestLog

from fec_dbsync import constants as fec_dbsync_cons
from fec_dbsync.constants import WORKFLOW_REPORT_URL_DICT, MIS_SHARED_PATH
from fec_dbsync.models import RoboIdtView, RoboRepayUplView
from math import sin, cos, sqrt, atan2, radians


def calculate_total_overdue_amount(robo_idt_view, robo_eit_obj=None, chat_object=None):
    no_of_overdue_terms = 0
    total_overdue_amount = 0
    total_paid_intallments = 0
    last_inst_obj = None
    for idt in robo_idt_view:
        if idt['DMD_AMT_SUM'] - idt['TOT_ADJ_AMT_SUM'] > 0:
            total_overdue_amount += idt['DMD_AMT_SUM'] - idt['TOT_ADJ_AMT_SUM']
            no_of_overdue_terms += 1
        else:
            total_paid_intallments += 1
        last_inst_obj = idt

    # Adding Penal Interest
    if robo_eit_obj:
        total_overdue_amount += (robo_eit_obj.PENAL_BOOKED_AMOUNT_DR -
                                 robo_eit_obj.PENAL_INTEREST_AMOUNT_DR)

    return total_overdue_amount, no_of_overdue_terms, total_paid_intallments, last_inst_obj


def calculate_overdue_terms(robo_idt_view):
    no_of_overdue_terms = 0
    for idt in robo_idt_view:
        if idt['DMD_AMT_SUM'] - idt['TOT_ADJ_AMT_SUM'] > 0:
            no_of_overdue_terms += 1

    return no_of_overdue_terms


def get_upcoming_due_date(robo_lrs_obj):
    for lrs in robo_lrs_obj:
        if lrs.FLOW_ID == 'EIDEM':
            return lrs.NEXT_DMD_DATE


def get_overdue_fee(chat_object):
    if chat_object and chat_object.CHARGE_TYPE == 'MISC2':
        return max(
            chat_object.USER_CALC_CHRGE_AMT, chat_object.SYS_CALC_CHRGE_AMT
        ) - chat_object.CHRGE_AMT_COLLECTED - chat_object.CHRGE_WAIVED


def get_monthly_installment(robo_lrs_obj):
    for lrs in robo_lrs_obj:
        if lrs.FLOW_ID == 'EIDEM':
            return lrs.FLOW_AMT


def get_repayments_info(idt_objects, business_date=None):
    payment_information = list()
    if business_date is None:
        business_date = datetime.date.today().strftime("%Y-%m-%d")

    for robo_idt_object in idt_objects:
        if robo_idt_object['DMD_EFF_DATE'].strftime("%Y-%m-%d") > business_date:
            pay_status = 'Unpaid'

            payment_information.append({
                'payment_date': robo_idt_object['DMD_EFF_DATE'],
                'payment_amount': robo_idt_object['DMD_AMT_SUM'] + fec_dbsync_cons.PROCESSING_FEE,
                'payment_status': pay_status
            })
            break
        else:
            pay_status = 'Overdue' if robo_idt_object['DMD_AMT_SUM'] - robo_idt_object['TOT_ADJ_AMT_SUM'] > 0 else \
                'Paid'

            payment_information.append({
                'payment_date': robo_idt_object['DMD_EFF_DATE'],
                'payment_amount': robo_idt_object['DMD_AMT_SUM'] + fec_dbsync_cons.PROCESSING_FEE,
                'payment_status': pay_status
            })

    return payment_information[::-1]


def convert_to_money(num="0"):
    l = []
    c = len(num)
    for i in range(int(math.ceil(c / 3))):
        l.append(num[max(0, c - 3):c])
        c -= 3
    return ','.join(reversed(l))


def format_number(num, percent=0):
    if num == "" or num == None or num == "NaN":
        return "-"
    if type(num) == float:
        if num == 0.0:
            return "-"
        num = "{0:.2f}".format(num)
    if type(num) == int:
        if num == 0:
            return "-"
        num = str(num)
    if type(num) == str and num.find('.') > -1:
        if float(num) == 0.0:
            return "-"
        num = "{0:.2f}".format(float(num))
    elif type(num) == str and num.find('.') == -1:
        if int(num) == 0:
            return "-"
        num = str(num)
    num_list = num.split('.')
    num_list[0] = convert_to_money(num_list[0])
    if percent:
        return '.'.join(num_list) + " %"
    return num_list[0]


def get_references_from_oracle(national_id, phone_number):
    SQL = "select * from TABLE(dw_bicc_cre.FN_REFL1_DETAILS('{}','{}'))"
    SQL = SQL.format(national_id, phone_number)
    print(SQL)
    connection = cx_Oracle.connect(
        "TUNGLT", "ThanhTung$", "172.27.1.7:1521/dwproddc", encoding="utf-8", nencoding="UTF-8")
    cursor = connection.cursor()
    cursor.execute(SQL)
    response = []
    for row in cursor:
        response.append(row)
    cursor.close()
    connection.close()
    return response


def get_data_from_oracle(national_id, phone_number):

    SQL = "select APP_DOMINO.FN_IS_DEDUP_CUST( '{}', '{}') from dual"
    SQL = SQL.format(national_id, phone_number)
    print(SQL)
    connection = cx_Oracle.connect(
        'APP_ROBO/CtgwRfrK9VaEJq3g@172.27.1.37:1521/cfappdb')
    cursor = connection.cursor()
    cursor.execute(SQL)
    for row in cursor:
        response = row
    cursor.close()
    connection.close()
    return response


def get_data_from_oracle_robo(phone_number, national_id, product_channel):

    SQL = "select DWROBOLENDING.FN_IS_DEDUPE_RL( '{}', '{}', '{}') from dual"
    SQL = SQL.format(phone_number, national_id, product_channel)
    print(SQL)
    connection = cx_Oracle.connect(
        'APP_ROBO/DWCtgwRfrK9VaEJq3g@172.27.1.8:1521/dwproddc')
    cursor = connection.cursor()
    cursor.execute(SQL)
    for row in cursor:
        response = row
    cursor.close()
    connection.close()
    return response


def get_data_from_workflow(url_identifier, filter=None):
    """
    get data from workflow url
    :param url_identifier:
    :param filter:
    :return:
    """
    login_url = settings.WORKFLOW_BASE_URL + "/login/credentials/"
    headers = {
        'Content-Type': "application/json",
        'Cache-Control': "no-cache"
    }
    data = {
        "username": os.environ.get('FLOWABLE_WORKFLOW_TEST_USER', 'super_user'),
        "password": os.environ.get('FLOWABLE_WORKFLOW_TEST_PASSWORD', 'P@ssw0rd^1223')
    }
    response = requests.post(login_url, data=json.dumps(data), headers=headers)
    response = response.json()
    if not response.get('status') == 200 or not response.get('data'):
        return {"status": False, "message": "Can't Login"}
    auth_token = response.get('data').get('bpm_token')
    workflow_url = WORKFLOW_REPORT_URL_DICT.get(url_identifier, None)
    if workflow_url is None:
        return {"status": False, "message": "URL mapping is not available."}
    url = settings.WORKFLOW_BASE_URL + "/mis/" + workflow_url
    if filter:
        query_params = urlencode(filter)
        url = url + '?' + query_params
    headers = {"Authorization": "Bearer " + auth_token}
    try:
        response = requests.get(url, headers=headers)
        response = response.json()
        if response["status"] == 200:
            return {"status": True, "response": response}
        else:
            return {"status": False, "message": "No Data Recieved from Flowable"}
    except:
        return {"status": False, "message": "Some Error Occured while Parsing Flowable Response"}


def mis_customer_journey_tracking_report(id, filters=None, emails=[]):
    """
    to generate mis report for customer journey tracking
    :return:
    """
    customerJourneyTracking = get_data_from_workflow(
        "customer_journey_tracking", filters)
    if customerJourneyTracking.get('status') == False or not customerJourneyTracking.get('status'):
        return None
    workbook = openpyxl.load_workbook(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'doc-template/customer_journey_tracking.xlsx'))
    response_data = customerJourneyTracking.get('response').get('data')
    work_sheet = workbook['Customer Journey Tracking']

    work_sheet[chr(68) + str(34)].value = filters.get("from",
                                                      "") if filters else ""
    work_sheet[chr(70) + str(34)].value = filters.get("to",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(35)].value = filters.get("product",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(36)
               ].value = filters.get("customertype", "") if filters else ""
    work_sheet[chr(68) + str(37)].value = filters.get("province",
                                                      "") if filters else ""
    work_sheet[chr(72) + str(34)].value = id.split("_")[0]

    row = 39
    for email in emails:
        work_sheet[chr(66) + str(row)].value = email
        row += 1

    row = 7
    update_list = [
        'TotalLeads',
        'SuccessfulApplications',
        'leadsInProgress',
        'BasicCustomerDetails',
        'DigitalQualification',
        'LoanDetails',
        'ApplicationReview',
        'CreditDecision',
        'Esign',
        'Disbursement',
        'Total',
    ]
    taskFields = response_data.get('taskFields')
    taskFields.append(response_data.get('totalLeads'))
    for i in update_list:
        data = next((item for item in taskFields if item['name'] == i))
        work_sheet[chr(68) + str(row)
                   ].value = format_number(data.get('ios').get('number'))
        work_sheet[chr(69) + str(row)
                   ].value = format_number(data.get('ios').get('amount'))
        work_sheet[chr(70) + str(row)
                   ].value = format_number(data.get('android').get('number'))
        work_sheet[chr(71) + str(row)
                   ].value = format_number(data.get('android').get('amount'))
        work_sheet[chr(72) + str(row)
                   ].value = format_number(data.get('web').get('number'))
        work_sheet[chr(73) + str(row)
                   ].value = format_number(data.get('web').get('amount'))
        work_sheet[chr(74) + str(row)
                   ].value = format_number(data.get('total').get('number'))
        work_sheet[chr(75) + str(row)
                   ].value = format_number(data.get('total').get('amount'))
        row += 1
        if row == 10:
            row += 1

    percent_update_list = [
        'PercentIos',
        'PercentAndroid',
        'PercentWeb',
        'PercentageBasicCustomerDetails',
        'PercentageDigitalQualification',
        'PercentageLoanDetails',
        'PercentageApplicationReview',
        'PercentageCreditDecision',
        'PercentageEsign',
        'PercentageDisbursement',
        'PercentageLeadsInProgress',
        'PercentageLeadConversion'
    ]
    for i in percent_update_list:
        data = next((item for item in response_data.get(
            'taskPercentFields') if item['name'] == i))
        work_sheet[chr(68) + str(row)
                   ].value = format_number(data.get('ios', 0), 1)
        work_sheet[chr(69) + str(row)
                   ].value = format_number(data.get('iosAmount', 0), 1)
        work_sheet[chr(70) + str(row)
                   ].value = format_number(data.get('android', 0), 1)
        work_sheet[chr(71) + str(row)
                   ].value = format_number(data.get('androidAmount', 0), 1)
        work_sheet[chr(72) + str(row)
                   ].value = format_number(data.get('web', 0), 1)
        work_sheet[chr(73) + str(row)
                   ].value = format_number(data.get('webAmount', 0), 1)
        work_sheet[chr(74) + str(row)
                   ].value = format_number(data.get('total', 0), 1)
        work_sheet[chr(75) + str(row)
                   ].value = format_number(data.get('totalAmount', 0), 1)
        row += 1
        if row == 22:
            row += 1
    os.makedirs(MIS_SHARED_PATH, exist_ok=True)
    sheet_name = MIS_SHARED_PATH + \
        str(id).format("customer_journey_tracking") + ".xlsx"
    workbook.save(sheet_name)
    return sheet_name


def mis_reject_loan_application_report(id, filters=None, emails=[]):
    """
    to generate reject loan application report
    :return:
    """
    rejectedApplications = get_data_from_workflow(
        "reject_loan_application", filters)
    if rejectedApplications.get('status') == False or not rejectedApplications.get('status'):
        return rejectedApplications
    response_data = rejectedApplications.get('response').get('data')
    workbook = openpyxl.load_workbook(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'doc-template/reject_loan_application.xlsx'))
    work_sheet = workbook['Rejected Loan Applications']

    work_sheet[chr(68) + str(28)].value = filters.get("from",
                                                      "") if filters else ""
    work_sheet[chr(69) + str(28)].value = filters.get("to",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(29)
               ].value = filters.get("customertype", "") if filters else ""
    work_sheet[chr(68) + str(30)].value = filters.get("province",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(31)
               ].value = filters.get("customerAgeFrom", "") if filters else ""
    work_sheet[chr(69) + str(31)
               ].value = filters.get("customerAgeTo", "") if filters else ""
    work_sheet[chr(68) + str(32)].value = filters.get("platform",
                                                      "") if filters else ""
    work_sheet[chr(70) + str(28)].value = id.split("_")[0]

    # row = 34
    # for email in emails:
    #     work_sheet[chr(66) + str(row)].value = email
    #     row += 1

    row = 7
    data = response_data.get('casesRejected')
    data['rejectedApplications'] = response_data.get('rejectedApplications')
    update_list = [
        'rejectedApplications',
        'blacklist',
        'pcb',
        'gcl',
        'creditEngine',
        'vmg',
        'xsell',
        'identityVerification',
        'cifNameMismatch',
        'selfCancel',
        'percentageBlacklist',
        'percentagePCB',
        'percentageGCL',
        'percentageCreditEngine',
        'percentageVMG',
        'percentageXsell',
        'percentageIdentityVerification',
        'percentageCifNameMismatch',
        'percentageSelfCancel'
    ]

    for i in update_list:
        work_sheet[chr(68) + str(row)].value = format_number(data[i].get('number', 0), 1) if i[
            :7] == "percent" else format_number(
            data[i].get('number', 0))
        work_sheet[chr(69) + str(row)].value = format_number(data[i].get('amount', 0), 1) if i[
            :7] == "percent" else format_number(
            data[i].get('amount', 0))
        row += 1
    os.makedirs(MIS_SHARED_PATH, exist_ok=True)
    sheet_name = MIS_SHARED_PATH + \
        str(id).format("reject_loan_application") + ".xlsx"
    workbook.save(sheet_name)
    return sheet_name


def mis_api_status_tracking_report(id, filters=None, emails=[]):
    """
    to generate api status tracking report
    :return:
    """
    workbook = openpyxl.load_workbook(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'doc-template/api_status_tracking.xlsx'))
    work_sheet = workbook['API Status Tracking']
    row = 12
    if filters and filters.get("to") and filters.get("from"):
        to = (datetime.datetime.strptime(filters.get("to"), '%Y-%m-%d').date() + datetime.timedelta(days=1)).strftime(
            '%Y-%m-%d')
        api_data = APIRequestLog.objects.filter(
            requested_at__range=[filters.get("from"), to]).order_by('path')
    elif filters and filters.get("to"):
        api_data = APIRequestLog.objects.filter(
            requested_at__lte=filters.get("to")).order_by('path')
    elif filters and filters.get("from"):
        api_data = APIRequestLog.objects.filter(
            requested_at__gte=filters.get("from")).order_by('path')
    else:
        api_data = APIRequestLog.objects.filter().order_by('path')
    api_dict = {}

    work_sheet[chr(68) + str(6)].value = filters.get("from",
                                                     "") if filters else ""
    work_sheet[chr(69) + str(6)].value = filters.get("to",
                                                     "") if filters else ""
    work_sheet[chr(68) + str(7)].value = filters.get("api",
                                                     "") if filters else ""
    work_sheet[chr(68) + str(8)].value = filters.get("service",
                                                     "") if filters else ""
    work_sheet[chr(70) + str(6)].value = id.split("_")[0]

    if api_data:
        total_success, total_errors, total = 0, 0, 0
        for i in api_data:
            if filters and filters.get('api', 0) and i.path and filters.get('api', "").upper() != i.path.split('/')[
                    -3].upper():
                continue
            total += 1
            if not i.errors:
                total_success += 1
                continue
            else:
                total_errors += 1
            try:
                error = ast.literal_eval(i.errors)
            except:
                error = {'reason': "Others"}
            if not api_dict.get(i.path):
                api_dict[i.path] = {}
            if not api_dict[i.path].get(error.get('reason')):
                api_dict[i.path][error.get('reason')] = 0
            api_dict[i.path] = {
                error.get('reason'): api_dict[i.path].get(error.get('reason')) + 1
            }
        work_sheet[chr(69) + str(row)].value = format_number(total_success)
        row += 1
        work_sheet[chr(69) + str(row)].value = format_number(total_errors)
        row += 1
        work_sheet[chr(69) + str(row)].value = format_number(total)
        row += 1
        work_sheet[chr(
            69) + str(row)].value = format_number(total_success * 100 / max(1, total), 1)
        row += 1
        work_sheet[chr(
            69) + str(row)].value = format_number(total_errors * 100 / max(1, total), 1)
        row += 1
        count = 1
        row = 21
        for api in api_dict.keys():
            work_sheet[chr(67) + str(row)].value = "API Service " + \
                str(count) + ": " + api.split('/')[-3].upper()
            count2 = 1
            for reason in api_dict[api].keys():
                work_sheet[chr(66) + str(row)].value = count
                work_sheet[chr(68) + str(row)].value = "Reason " + \
                    str(count2) + ": " + reason
                work_sheet[chr(69) + str(row)
                           ].value = format_number(api_dict[api][reason])
                count2 += 1
                row += 1
                count += 1
        work_sheet[chr(66) + str(row)].value = "Total No. of Failed Call"
        work_sheet[chr(69) + str(row)].value = format_number(total_errors)
    os.makedirs(MIS_SHARED_PATH, exist_ok=True)
    sheet_name = MIS_SHARED_PATH + \
        str(id).format("api_status_tracking") + ".xlsx"
    workbook.save(sheet_name)
    return sheet_name


def mis_customer_portfolio_report(id, filters=None, emails=[]):
    """
    to generate custgomer portfolio report
    :return:
    """
    workbook = openpyxl.load_workbook(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'doc-template/customer_portfolio.xlsx'))
    work_sheet = workbook['Customer Portfolio']
    customerPortfolioData = get_data_from_workflow(
        "customer_portfolio", filters)
    if customerPortfolioData.get('status') == False or not customerPortfolioData.get('status'):
        return customerPortfolioData
    response_data = customerPortfolioData.get('response').get('data')

    work_sheet[chr(67) + str(6)].value = filters.get("from",
                                                     "") if filters else ""
    work_sheet[chr(68) + str(6)].value = filters.get("to",
                                                     "") if filters else ""
    work_sheet[chr(67) + str(8)].value = filters.get("loanStatus",
                                                     "") if filters else ""
    work_sheet[chr(67) + str(9)].value = filters.get("customerAgeFrom",
                                                     "") if filters else ""
    work_sheet[chr(68) + str(9)].value = filters.get("customerAgeTo",
                                                     "") if filters else ""
    work_sheet[chr(69) + str(6)].value = id.split("_")[0]

    row = 13
    count = 0
    total = [0, 0, 0, 0, 0, 0, 0, 0, 0]
    for data in response_data.keys():
        count += 1
        work_sheet[chr(65) + str(row)].value = count
        work_sheet[chr(66) + str(row)].value = data
        work_sheet[chr(67) + str(row)].value = format_number(
            response_data[data]['CustomerType']['Married'])
        total[0] += response_data[data]['CustomerType']['Married']
        work_sheet[chr(68) + str(row)].value = format_number(
            response_data[data]['CustomerType']['Single'])
        total[1] += response_data[data]['CustomerType']['Single']
        work_sheet[chr(69) + str(row)].value = format_number(
            response_data[data]['CustomerType']['Widow'])
        total[2] += response_data[data]['CustomerType']['Widow']
        work_sheet[chr(70) + str(row)].value = format_number(
            response_data[data]['CustomerType']['Divorced'])
        total[3] += response_data[data]['CustomerType']['Divorced']
        work_sheet[chr(71) + str(row)].value = format_number(
            response_data[data]['CustomerType']['Others'])
        total[4] += response_data[data]['CustomerType']['Others']
        work_sheet[chr(72) + str(row)].value = format_number(
            response_data[data]['CustomerEmploymentStatus']['Employed'])
        total[5] += response_data[data]['CustomerEmploymentStatus']['Employed']
        work_sheet[chr(73) + str(row)].value = format_number(
            response_data[data]['CustomerEmploymentStatus']['Self-Employed'])
        total[6] += response_data[data]['CustomerEmploymentStatus']['Self-Employed']
        work_sheet[chr(74) + str(row)].value = format_number(
            response_data[data]['CustomerBackground']['NTB'])
        total[7] += response_data[data]['CustomerBackground']['NTB']
        work_sheet[chr(75) + str(row)].value = format_number(
            response_data[data]['CustomerBackground']['Existing'])
        total[8] += response_data[data]['CustomerBackground']['Existing']
        row += 1

    # work_sheet[chr(65) + str(row)].value = count + 1
    work_sheet[chr(66) + str(row)].value = "Total"
    for i in range(len(total)):
        work_sheet[chr(67 + i) + str(row)].value = format_number(total[i])

    # row+=2
    # if emails and type(emails)==list:
    #     for email in emails:
    #         work_sheet[chr(66) + str(row)].value = email
    #         row += 1

    os.makedirs(MIS_SHARED_PATH, exist_ok=True)
    sheet_name = MIS_SHARED_PATH + \
        str(id).format("customer_portfolio") + ".xlsx"
    workbook.save(sheet_name)
    return sheet_name


def mis_credit_card_fi(id, filters=None, emails=[]):
    """
    FI Report for Credit Card
    :param id:
    :param filters:
    :param emails:
    :return:
    """
    workbook = openpyxl.load_workbook(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'doc-template/credit_card_fi.xlsx'))
    work_sheet = workbook['Credit Card Field Investigation']
    credit_card_fi_data = get_data_from_workflow("credit_card_fi", filters)
    if credit_card_fi_data.get('status') == False or not credit_card_fi_data.get('status'):
        return credit_card_fi_data
    response_data = credit_card_fi_data.get('response').get('data')
    row_no = 2
    for row in response_data.get('FIRows', []):
        work_sheet[chr(65) + str(row_no)].value = row.get("customerName", "")
        work_sheet[chr(66) + str(row_no)].value = row.get("phoneNumber", "")
        work_sheet[chr(67) + str(row_no)].value = row.get("nationalID", "")
        work_sheet[chr(68) + str(row_no)].value = row.get("placeOfIssue", "")
        work_sheet[chr(69) + str(row_no)].value = row.get("customerType", "")
        work_sheet[chr(70) + str(row_no)].value = row.get("CIFNo", "")
        work_sheet[chr(71) + str(row_no)].value = row.get("age", "")
        work_sheet[chr(72) + str(row_no)].value = row.get("sex", "")
        work_sheet[chr(73) + str(row_no)].value = row.get("DOB", "")
        work_sheet[chr(74) + str(row_no)].value = row.get("email", "")
        work_sheet[chr(75) + str(row_no)
                   ].value = row.get("temporaryAddress", "")
        work_sheet[chr(76) + str(row_no)
                   ].value = row.get("permanentAddress", "")
        work_sheet[chr(77) + str(row_no)].value = row.get("companyAddress", "")
        work_sheet[chr(78) + str(row_no)
                   ].value = row.get("applicationDate", "")
        work_sheet[chr(79) + str(row_no)
                   ].value = row.get("applicationStartDate", "")
        work_sheet[chr(80) + str(row_no)].value = row.get("creditLimit", "")
        work_sheet[chr(81) + str(row_no)].value = row.get("productCode", "")
        work_sheet[chr(82) + str(row_no)].value = row.get("schemeCode", "")
        work_sheet[chr(83) + str(row_no)
                   ].value = row.get("cardApplicationID", "")
        work_sheet[chr(84) + str(row_no)
                   ].value = row.get("applicationStatus", "")
        work_sheet[chr(86) + str(row_no)].value = row.get("addressChange", "")
        row_no += 1

    os.makedirs(MIS_SHARED_PATH, exist_ok=True)
    sheet_name = MIS_SHARED_PATH + str(id).format("credit_card_fi") + ".xlsx"
    workbook.save(sheet_name)
    return sheet_name


def mis_disbursement_status_tracking_report(id, filters=None, emails=[]):
    """
    to generate disbursement status tracking report
    :return:
    """
    workbook = openpyxl.load_workbook(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'doc-template/disbursement_status_tracking.xlsx'))
    work_sheet = workbook['Disbursement Status Tracking']
    disbursementStatusTracking = get_data_from_workflow(
        "disbursement_status_tracking", filters)
    if disbursementStatusTracking.get('status') == False or not disbursementStatusTracking.get('status'):
        return disbursementStatusTracking
    response_data = disbursementStatusTracking.get('response').get('data')

    work_sheet[chr(68) + str(31)].value = filters.get("from",
                                                      "") if filters else ""
    work_sheet[chr(69) + str(31)].value = filters.get("to",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(32)].value = filters.get("product",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(33)
               ].value = filters.get("customertype", "") if filters else ""
    work_sheet[chr(68) + str(34)].value = filters.get("province",
                                                      "") if filters else ""
    work_sheet[chr(70) + str(31)].value = id.split("_")[0]

    row = 36
    for email in emails:
        work_sheet[chr(66) + str(row)].value = email
        row += 1

    row = 6
    update_list = [
        'approvedLoans',
        'successfulDisbursement',
        'failedDisbursement',
        'pendingLoans',
        'cancelledDisbursement',
        'cashDisbursement',
        'transferDisbursement',
        'percentSuccessfulDisbursement',
        'percentUnsuccessfulDisbursement',
        'percentcashDisbursement',
        'percenttransferDisbursement'
    ]
    for line in update_list:
        work_sheet[chr(68) + str(row)].value = format_number(
            response_data[line].get('total', response_data[line]).get('number', 0), 1) if line[
            :7] == "percent" else format_number(
            response_data[line].get('total', response_data[line]).get('number'))
        work_sheet[chr(71) + str(row)].value = format_number(
            response_data[line].get('total', response_data[line]).get('amount', 0), 1) if line[
            :7] == "percent" else format_number(
            response_data[line].get('total', response_data[line]).get('amount'))
        row += 1

    update_list = [
        'approvedLoans',
        'successfulDisbursement',
        'failedDisbursement',
        'pendingLoans',
        'percentSuccessfulChannel',
        'percentUnsuccessfulChannel'
    ]
    row = 22
    for line in update_list:
        work_sheet[chr(68) + str(row)].value = format_number(response_data[line].get('VN').get('number', 0), 1) if line[
            :7] == "percent" else format_number(
            response_data[line].get('VN').get(
                'number'))
        work_sheet[chr(69) + str(row)].value = format_number(response_data[line].get('VN').get('amount', 0), 1) if line[
            :7] == "percent" else format_number(
            response_data[line].get('VN').get(
                'amount'))
        work_sheet[chr(74) + str(row)].value = format_number(response_data[line].get('BI').get('number', 0), 1) if line[
            :7] == "percent" else format_number(
            response_data[line].get('BI').get(
                'number'))
        work_sheet[chr(75) + str(row)].value = format_number(response_data[line].get('BI').get('amount', 0), 1) if line[
            :7] == "percent" else format_number(
            response_data[line].get('BI').get(
                'amount'))
        work_sheet[chr(86) + str(row)].value = format_number(response_data[line].get('AT').get('number', 0), 1) if line[
            :7] == "percent" else format_number(
            response_data[line].get('AT').get(
                'number'))
        work_sheet[chr(87) + str(row)].value = format_number(response_data[line].get('AT').get('amount', 0), 1) if line[
            :7] == "percent" else format_number(
            response_data[line].get('AT').get(
                'amount'))
        # work_sheet[chr(89) + str(row)].value = format_number(response_data[line].get('SA').get('number', 0), 1) if line[
        #                                                                                                            :7] == "percent" else format_number(
        #     response_data[line].get('SA').get(
        #         'number'))
        # work_sheet[chr(90) + str(row)].value = format_number(response_data[line].get('SA').get('amount', 0), 1) if line[
        #                                                                                                            :7] == "percent" else format_number(
        #     response_data[line].get('SA').get(
        #         'amount'))
        # work_sheet[chr(65) + chr(66) + str(row)].value = format_number(response_data[line].get('AG').get('number', 0),
        #                                                                1) if line[:7] == "percent" else format_number(
        #     response_data[line].get(
        #         'AG').get(
        #         'number'))
        # work_sheet[chr(65) + chr(67) + str(row)].value = format_number(response_data[line].get('AG').get('amount', 0),
        #                                                                1) if line[:7] == "percent" else format_number(
        #     response_data[line].get(
        #         'AG').get(
        #         'amount'))
        row += 1
    os.makedirs(MIS_SHARED_PATH, exist_ok=True)
    sheet_name = MIS_SHARED_PATH + \
        str(id).format("disbursement_status_tracking") + ".xlsx"
    workbook.save(sheet_name)
    return sheet_name


def mis_robo_sales_report(id, filters=None, emails=[]):
    """
    to generate robo sales report
    :return:
    """
    disbursementStatusTracking = get_data_from_workflow(
        "disbursement_status_tracking", filters)
    if disbursementStatusTracking.get('status') == False or not disbursementStatusTracking.get('status'):
        return disbursementStatusTracking
    disbursementStatusTracking = disbursementStatusTracking.get('response')
    rejectedApplications = get_data_from_workflow(
        "reject_loan_application", filters)
    if rejectedApplications.get('status') == False or not rejectedApplications.get('status'):
        return rejectedApplications
    rejectedApplications = rejectedApplications.get('response').get('data')
    workbook = openpyxl.load_workbook(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'doc-template/robo_sales.xlsx'))
    work_sheet = workbook['Robo Sales']

    work_sheet[chr(68) + str(18)].value = filters.get("from",
                                                      "") if filters else ""
    work_sheet[chr(69) + str(18)].value = filters.get("to",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(19)].value = filters.get("product",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(20)
               ].value = filters.get("customertype", "") if filters else ""
    work_sheet[chr(68) + str(21)].value = filters.get("province",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(22)].value = filters.get("platform",
                                                      "") if filters else ""
    work_sheet[chr(70) + str(18)].value = id.split("_")[0]

    # row = 25
    # for email in emails:
    #     work_sheet[chr(66) + str(row)].value = email
    #     row += 1

    work_sheet[chr(68) + str(7)].value = format_number(
        rejectedApplications['totalLeads']['total']['number'])
    work_sheet[chr(69) + str(7)].value = format_number(
        rejectedApplications['totalLeads']['total']['amount'])
    work_sheet[chr(68) + str(8)].value = format_number(
        disbursementStatusTracking['data']['approvedLoans']['total']['number'] +
        disbursementStatusTracking['data']['failedDisbursement']['total']['number'] +
        disbursementStatusTracking['data']['pendingLoans']['total']['number'] +
        disbursementStatusTracking['data']['cancelledDisbursement']['total'][
            'number'] + rejectedApplications['casesRejected']['creditEngine']['number'])
    work_sheet[chr(69) + str(8)].value = format_number(
        disbursementStatusTracking['data']['approvedLoans']['total']['amount'] +
        disbursementStatusTracking['data']['failedDisbursement']['total']['amount'] +
        disbursementStatusTracking['data']['pendingLoans']['total']['amount'] +
        disbursementStatusTracking['data']['cancelledDisbursement']['total'][
            'amount'] + rejectedApplications['casesRejected']['creditEngine']['amount'])
    work_sheet[chr(68) + str(9)].value = "-"
    work_sheet[chr(69) + str(9)].value = "-"
    work_sheet[chr(68) + str(10)].value = format_number(disbursementStatusTracking['data']['insuredLoans']['total'][
        'number'])
    work_sheet[chr(69) + str(10)].value = format_number(disbursementStatusTracking['data']['insuredLoans']['total'][
        'amount'])
    work_sheet[chr(68) + str(11)].value = format_number(
        disbursementStatusTracking['data']['expiredDisbursement']['total']['number'])
    work_sheet[chr(69) + str(11)].value = format_number(
        disbursementStatusTracking['data']['expiredDisbursement']['total']['amount'])
    work_sheet[chr(68) + str(12)].value = format_number(
        disbursementStatusTracking['data']['closedDisbursement']['total'][
            'number'])
    work_sheet[chr(69) + str(12)].value = format_number(
        disbursementStatusTracking['data']['closedDisbursement']['total'][
            'amount'])
    work_sheet[chr(68) + str(13)].value = format_number(
        (disbursementStatusTracking['data']['approvedLoans']['total']['number'] +
         disbursementStatusTracking['data']['failedDisbursement']['total']['number'] +
         disbursementStatusTracking['data']['pendingLoans']['total']['number'] +
         disbursementStatusTracking['data']['cancelledDisbursement']['total'][
             'number'] + rejectedApplications['casesRejected']['creditEngine']['number']) * 100 / max(1, int(
                 rejectedApplications['totalLeads']['total']['number'])), 1)
    work_sheet[chr(69) + str(13)].value = format_number(
        (disbursementStatusTracking['data']['approvedLoans']['total']['amount'] +
         disbursementStatusTracking['data']['failedDisbursement']['total']['amount'] +
         disbursementStatusTracking['data']['pendingLoans']['total']['amount'] +
         disbursementStatusTracking['data']['cancelledDisbursement']['total'][
             'amount'] + rejectedApplications['casesRejected']['creditEngine']['amount']) * 100 / max(1, int(
                 rejectedApplications['totalLeads']['total']['amount'])), 1)
    work_sheet[chr(68) + str(14)].value = format_number(
        disbursementStatusTracking['data']['approvedLoans']['total']['amount'] /
        max(1, disbursementStatusTracking['data']['approvedLoans']['total']['number']))
    work_sheet[chr(68) + str(15)].value = format_number(
        disbursementStatusTracking['data']['insuredLoans']['total']['amount'] /
        max(1, disbursementStatusTracking['data']['approvedLoans']['total']['number']))
    os.makedirs(MIS_SHARED_PATH, exist_ok=True)
    sheet_name = MIS_SHARED_PATH + str(id).format("robo_sales") + ".xlsx"
    workbook.save(sheet_name)
    return sheet_name


def mis_loan_snapshot_active_report(id, filters=None, emails=[]):
    """
    to generate loan snapshot active report
    :return:
    """
    loanPortfolioSnapshot = get_data_from_workflow(
        "loan_portfolio_snapshot", filters)
    if loanPortfolioSnapshot.get('status') == False or not loanPortfolioSnapshot.get('status'):
        return loanPortfolioSnapshot
    workbook = openpyxl.load_workbook(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'doc-template/loan_snapshot_active.xlsx'))
    work_sheet = workbook["Loan Snapshot Active"]

    work_sheet[chr(68) + str(6)].value = filters.get("from",
                                                     "") if filters else ""
    work_sheet[chr(69) + str(6)].value = filters.get("to",
                                                     "") if filters else ""
    work_sheet[chr(68) + str(7)].value = filters.get("province",
                                                     "") if filters else ""
    work_sheet[chr(68) + str(8)].value = filters.get("platform",
                                                     "") if filters else ""
    work_sheet[chr(70) + str(6)].value = id.split("_")[0]

    response_data = loanPortfolioSnapshot.get('response')['data']
    row = 11
    count1 = 1
    for line1 in response_data.keys():
        total = [0] * 4
        total_row = row
        count2 = 1
        work_sheet[chr(66) + str(row)].value = format_number(count1)
        work_sheet[chr(67) + str(row)].value = line1
        row += 1
        for line2 in response_data[line1].keys():
            work_sheet[chr(66) + str(row)].value = str(count1) + \
                "." + str(count2)
            work_sheet[chr(67) + str(row)].value = line2
            work_sheet[chr(68) + str(row)].value = format_number(
                (response_data[line1][line2].get('Employed')['number'] if
                 response_data[line1][line2].get('Employed') else 0) + (
                    response_data[line1][line2].get('Salaried')['number'] if
                    response_data[line1][line2].get('Salaried') else 0))
            total[0] += (response_data[line1][line2].get('Employed')['number'] if
                         response_data[line1][line2].get('Employed') else 0) + (
                response_data[line1][line2].get('Salaried')['number'] if
                response_data[line1][line2].get('Salaried') else 0)
            work_sheet[chr(69) + str(row)].value = format_number(
                (response_data[line1][line2].get('Employed')['amount'] if
                 response_data[line1][line2].get('Employed') else 0) + (
                    response_data[line1][line2].get('Salaried')['amount'] if
                    response_data[line1][line2].get('Salaried') else 0))
            total[1] += (response_data[line1][line2].get('Employed')['amount'] if
                         response_data[line1][line2].get('Employed') else 0) + (
                response_data[line1][line2].get('Salaried')['amount'] if
                response_data[line1][line2].get('Salaried') else 0)
            work_sheet[chr(70) + str(row)].value = format_number(
                response_data[line1][line2].get('Self-employed', {}).get('number'))
            total[2] += response_data[line1][line2].get(
                'Self-employed', {}).get('number', 0)
            work_sheet[chr(71) + str(row)].value = format_number(
                response_data[line1][line2].get('Self-employed', {}).get('amount'))
            total[3] += response_data[line1][line2].get(
                'Self-employed', {}).get('amount', 0)
            count2 += 1
            row += 1
        work_sheet[chr(68) + str(total_row)].value = format_number(total[0])
        work_sheet[chr(69) + str(total_row)].value = format_number(total[1])
        work_sheet[chr(70) + str(total_row)].value = format_number(total[2])
        work_sheet[chr(71) + str(total_row)].value = format_number(total[3])
        count1 += 1

    row += 2
    for email in emails:
        work_sheet[chr(66) + str(row)].value = email
        row += 1

    os.makedirs(MIS_SHARED_PATH, exist_ok=True)
    sheet_name = MIS_SHARED_PATH + \
        str(id).format("loan_snapshot_active") + ".xlsx"
    workbook.save(sheet_name)
    return sheet_name


def mis_product_performance_report(id, filters=None, emails=[]):
    """
    to generate product performance report
    :return:
    """
    productPerformanceReport = get_data_from_workflow(
        "product_performance_report", filters)
    if productPerformanceReport.get('status') == False or not productPerformanceReport.get('status'):
        return productPerformanceReport
    workbook = openpyxl.load_workbook(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'doc-template/product_performance.xlsx'))
    work_sheet = workbook["Product Performance"]

    work_sheet[chr(68) + str(37)].value = filters.get("from",
                                                      "") if filters else ""
    work_sheet[chr(69) + str(37)].value = filters.get("to",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(38)].value = filters.get("product",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(39)].value = filters.get("scheme",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(40)
               ].value = filters.get("customertype", "") if filters else ""
    work_sheet[chr(68) + str(41)
               ].value = filters.get("customerStatus", "") if filters else ""
    work_sheet[chr(70) + str(37)].value = id.split("_")[0]

    # row = 44
    # for email in emails:
    #     work_sheet[chr(66) + str(row)].value = email
    #     row += 1

    response_data = productPerformanceReport.get('response')['data']
    row = 7
    update_list = [
        'loansApplied',
        'loansApproved',
        'approvedVMG',
        'approvedWithoutVMG',
        'approvedMRC',
        'approvedEVN',
        'approvedXSellDB',
        'ApprovedrequestedGreater',
        # 'AppliedrequestedGreater',
        'percentageApproved',
        'percentageXSellDB',
        'percentageVMG',
        'percentageWithoutVMG',
        'percentageMRC',
        'percentageEVN'
    ]
    skip_rows = [15, 16, 17, 18, 19, 20, 22, 23, 24, 25, 26, 27]
    for line in update_list:
        while row in skip_rows:
            row += 1
        percent_sign = 1 if line[:7] == "percent" else 0
        work_sheet[chr(
            68) + str(row)].value = format_number(response_data[line]['number'], percent_sign)
        work_sheet[chr(
            69) + str(row)].value = format_number(response_data[line]['amount'], percent_sign)
        work_sheet[chr(70) + str(row)].value = format_number(
            response_data[line]['avgtenure'], percent_sign)
        work_sheet[chr(71) + str(row)].value = format_number(
            response_data[line]['avgAmount'], percent_sign)
        row += 1
    os.makedirs(MIS_SHARED_PATH, exist_ok=True)
    sheet_name = MIS_SHARED_PATH + \
        str(id).format("product_performance") + ".xlsx"
    workbook.save(sheet_name)
    return sheet_name


def mis_loan_health_check(id, filters=None, emails=[]):
    """
    to generate loan health check report
    :return:
    """
    loanHealthCheck = get_data_from_workflow("loan_health_check", filters)
    if loanHealthCheck.get('status') == False or not loanHealthCheck.get('status'):
        return loanHealthCheck
    workbook = openpyxl.load_workbook(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'doc-template/loan_health_check.xlsx'))
    work_sheet = workbook["Loan Health Check"]

    # work_sheet[chr(68) + str(37)].value = filters.get("from", "") if filters else ""
    # work_sheet[chr(69) + str(37)].value = filters.get("to", "") if filters else ""
    work_sheet[chr(68) + str(5)].value = filters.get("product",
                                                     "") if filters else ""
    work_sheet[chr(68) + str(6)].value = filters.get("scheme",
                                                     "") if filters else ""
    work_sheet[chr(68) + str(7)].value = filters.get("customertype",
                                                     "") if filters else ""
    work_sheet[chr(68) + str(8)].value = filters.get("tenure",
                                                     "") if filters else ""
    work_sheet[chr(68) + str(9)].value = filters.get("province",
                                                     "") if filters else ""
    work_sheet[chr(70) + str(4)].value = id.split("_")[0]

    # row = 44
    # for email in emails:
    #     work_sheet[chr(66) + str(row)].value = email
    #     row += 1

    response_data = loanHealthCheck.get('response')['data']
    row = 11
    buckets = 13
    for line in range(buckets):
        work_sheet[chr(68) + str(row)].value = format_number(
            response_data.get("bucket" + str(line), {}).get('number', 0))
        work_sheet[chr(69) + str(row)].value = format_number(
            response_data.get("bucket" + str(line), {}).get('amount', 0))
        # print(format_number(float(response_data.get("bucket" + str(line), {}).get('overdue', 0))))
        work_sheet[chr(70) + str(row)].value = format_number(
            float(response_data.get("bucket" + str(line), {}).get('overdue', 0)) * 100, 1)
        row += 1

    work_sheet[chr(68) + str(row)
               ].value = format_number(response_data["totalOverdue"]['number'])
    work_sheet[chr(69) + str(row)
               ].value = format_number(response_data["totalOverdue"]['amount'])
    work_sheet[chr(70) + str(row)].value = format_number(
        float(response_data["totalOverdue"]['overdue']) * 100, 1)
    row += 1

    work_sheet[chr(68) + str(row)].value = format_number(
        response_data["totalOutstandingloans"]['number'])
    work_sheet[chr(69) + str(row)].value = format_number(
        response_data["totalOutstandingloans"]['amount'])
    work_sheet[chr(70) + str(row)].value = format_number(
        response_data["totalOutstandingloans"]['overdue'] * 100, 1)

    os.makedirs(MIS_SHARED_PATH, exist_ok=True)
    sheet_name = MIS_SHARED_PATH + \
        str(id).format("loan_health_check") + ".xlsx"
    workbook.save(sheet_name)
    return sheet_name


def mis_loan_status_tracking_report(id, filters=None, emails=[]):
    """
    to generate loan status tracking
    :return:
    """
    disbursementStatusTracking = get_data_from_workflow(
        "disbursement_status_tracking", filters)
    if disbursementStatusTracking.get('status') == False or not disbursementStatusTracking.get('status'):
        return disbursementStatusTracking
    disbursementStatusTracking = disbursementStatusTracking.get('response')
    rejectedLoanApplication = get_data_from_workflow(
        "reject_loan_application", filters)
    if rejectedLoanApplication.get('status') == False or not rejectedLoanApplication.get('status'):
        return rejectedLoanApplication
    rejectedLoanApplication = rejectedLoanApplication.get('response')
    workbook = openpyxl.load_workbook(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'doc-template/loan_status_tracking.xlsx'))
    work_sheet = workbook["Loan Status Tracking"]

    work_sheet[chr(68) + str(28)].value = filters.get("from",
                                                      "") if filters else ""
    work_sheet[chr(69) + str(28)].value = filters.get("to",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(29)].value = filters.get("product",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(30)
               ].value = filters.get("customertype", "") if filters else ""
    work_sheet[chr(68) + str(31)].value = filters.get("province",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(32)].value = filters.get("loanStatus",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(33)].value = filters.get("platform",
                                                      "") if filters else ""
    work_sheet[chr(70) + str(28)].value = id.split("_")[0]

    # row = 35
    # for email in emails:
    #     work_sheet[chr(66) + str(row)].value = email
    #     row += 1

    work_sheet[chr(68) + str(7)].value = format_number(
        rejectedLoanApplication['data']['totalLeads']['total']['number'])
    work_sheet[chr(69) + str(7)].value = format_number(
        rejectedLoanApplication['data']['totalLeads']['total']['amount'])

    work_sheet[chr(68) + str(9)].value = format_number(
        disbursementStatusTracking['data']['approvedLoans']['total']['number'])
    work_sheet[chr(69) + str(9)].value = format_number(
        disbursementStatusTracking['data']['approvedLoans']['total']['amount'])
    work_sheet[chr(68) + str(10)].value = format_number(
        rejectedLoanApplication['data']['rejectedApplications']['number'])
    work_sheet[chr(69) + str(10)].value = format_number(
        rejectedLoanApplication['data']['rejectedApplications']['amount'])
    work_sheet[chr(68) + str(11)].value = format_number(
        disbursementStatusTracking['data']['successfulDisbursement']['total'][
            'number'])
    work_sheet[chr(69) + str(11)].value = format_number(
        disbursementStatusTracking['data']['successfulDisbursement']['total'][
            'amount'])
    work_sheet[chr(68) + str(12)].value = "-"
    work_sheet[chr(69) + str(12)].value = "-"
    work_sheet[chr(68) + str(13)].value = format_number(
        disbursementStatusTracking['data']['insuredLoans']['total']['number'])
    work_sheet[chr(69) + str(13)].value = format_number(
        disbursementStatusTracking['data']['insuredLoans']['total']['amount'])
    work_sheet[chr(68) + str(14)].value = format_number(
        disbursementStatusTracking['data']['nonInsuredLoans']['total']['number'])
    work_sheet[chr(69) + str(14)].value = format_number(
        disbursementStatusTracking['data']['nonInsuredLoans']['total']['amount'])
    work_sheet[chr(68) + str(15)].value = format_number(
        disbursementStatusTracking['data']['expiredDisbursement']['total']['number'])
    work_sheet[chr(69) + str(15)].value = format_number(
        disbursementStatusTracking['data']['expiredDisbursement']['total']['amount'])
    work_sheet[chr(68) + str(16)].value = format_number(
        disbursementStatusTracking['data']['closedDisbursement']['total']['number'])
    work_sheet[chr(69) + str(16)].value = format_number(
        disbursementStatusTracking['data']['closedDisbursement']['total']['amount'])

    work_sheet[chr(68) + str(17)].value = format_number(
        disbursementStatusTracking['data']['approvedLoans']['total']['number'] * 100 /
        max(1, rejectedLoanApplication['data']['totalLeads']['total']['number']), 1)
    work_sheet[chr(69) + str(17)].value = format_number(
        disbursementStatusTracking['data']['approvedLoans']['total']['amount'] * 100 /
        max(1, rejectedLoanApplication['data']['totalLeads']['total']['amount']), 1)
    work_sheet[chr(68) + str(18)].value = format_number(
        rejectedLoanApplication['data']['rejectedApplications']['number'] * 100 /
        max(1, rejectedLoanApplication['data']['totalLeads']['total']['number']), 1)
    work_sheet[chr(69) + str(18)].value = format_number(
        rejectedLoanApplication['data']['rejectedApplications']['amount'] * 100 /
        max(1, rejectedLoanApplication['data']['totalLeads']['total']['amount']), 1)
    work_sheet[chr(68) + str(19)].value = format_number(
        disbursementStatusTracking['data']['successfulDisbursement']['total']['number'] * 100 /
        max(1, disbursementStatusTracking['data']['approvedLoans']['total']['number']), 1)
    work_sheet[chr(69) + str(19)].value = format_number(
        disbursementStatusTracking['data']['successfulDisbursement']['total']['amount'] * 100 /
        max(1, disbursementStatusTracking['data']['approvedLoans']['total']['amount']), 1)
    work_sheet[chr(68) + str(20)].value = "-"
    work_sheet[chr(69) + str(20)].value = "-"
    work_sheet[chr(68) + str(21)].value = format_number(
        disbursementStatusTracking['data']['insuredLoans']['total']['number'] * 100 /
        max(1, rejectedLoanApplication['data']['totalLeads']['total']['number']), 1)
    work_sheet[chr(69) + str(21)].value = format_number(
        disbursementStatusTracking['data']['insuredLoans']['total']['amount'] * 100 /
        max(1, rejectedLoanApplication['data']['totalLeads']['total']['amount']), 1)
    work_sheet[chr(68) + str(22)].value = format_number(
        (disbursementStatusTracking['data']['nonInsuredLoans']['total']['number']) * 100 /
        max(1, disbursementStatusTracking['data']['appliedLoans']['total']['number']), 1)
    work_sheet[chr(69) + str(22)].value = format_number(
        (disbursementStatusTracking['data']['nonInsuredLoans']['total']['amount']) * 100 /
        max(1, disbursementStatusTracking['data']['appliedLoans']['total']['amount']), 1)
    work_sheet[chr(68) + str(23)].value = format_number(
        disbursementStatusTracking['data']['expiredDisbursement']['total']['number'] * 100 /
        max(1, disbursementStatusTracking['data']['approvedLoans']['total']['number']), 1)
    work_sheet[chr(69) + str(23)].value = format_number(
        disbursementStatusTracking['data']['expiredDisbursement']['total']['amount'] * 100 /
        max(1, disbursementStatusTracking['data']['approvedLoans']['total']['amount']), 1)
    work_sheet[chr(68) + str(24)].value = format_number(
        disbursementStatusTracking['data']['closedDisbursement']['total']['number'] * 100 /
        max(1, disbursementStatusTracking['data']['approvedLoans']['total']['number']), 1)
    work_sheet[chr(69) + str(24)].value = format_number(
        disbursementStatusTracking['data']['closedDisbursement']['total']['amount'] * 100 /
        max(1, disbursementStatusTracking['data']['approvedLoans']['total']['amount']), 1)
    work_sheet[chr(68) + str(25)].value = format_number(
        disbursementStatusTracking['data']['approvedLoans']['total']['amount'] /
        max(1, disbursementStatusTracking['data']['approvedLoans']['total']['number']))

    os.makedirs(MIS_SHARED_PATH, exist_ok=True)
    sheet_name = MIS_SHARED_PATH + \
        str(id).format("loan_status_tracking") + ".xlsx"
    workbook.save(sheet_name)
    return sheet_name


def mis_customer_service_tracking_report(id, filters=None, emails=[]):
    """
    to generate customer service tracking report
    :return:
    """
    customerServiceTracking = get_data_from_workflow(
        "customer_service_tracking", filters)
    if customerServiceTracking.get('status') == False or not customerServiceTracking.get('status'):
        return customerServiceTracking
    customerServiceTracking = customerServiceTracking.get('response')
    workbook = openpyxl.load_workbook(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'doc-template/customer_service_tracking.xlsx'))
    work_sheet = workbook['Customer Service Tracking']

    work_sheet[chr(68) + str(18)].value = filters.get("from",
                                                      "") if filters else ""
    work_sheet[chr(69) + str(18)].value = filters.get("to",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(19)].value = filters.get("category",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(20)].value = filters.get("product",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(21)].value = filters.get("scheme",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(22)].value = filters.get("agent",
                                                      "") if filters else ""
    work_sheet[chr(70) + str(18)].value = id.split("_")[0]

    row = 24
    for email in emails:
        work_sheet[chr(66) + str(row)].value = email
        row += 1

    response_data = customerServiceTracking.get('data')

    row = 6
    update_list = [
        'noCalls',
        'noServiceRequestRaised',
        'noServiceRequestPending',
        'noQueriesRaised',
        'noQueriesPending',
        'noApproved',
        'noRejected'
    ]
    for line in update_list:
        work_sheet[chr(68) + str(row)].value = format_number(response_data.get(line, 0), 1) \
            if line[:7] == "percent" else format_number(response_data.get(line, 0))
        row += 1
    os.makedirs(MIS_SHARED_PATH, exist_ok=True)
    sheet_name = MIS_SHARED_PATH + \
        str(id).format("customer_service_tracking") + ".xlsx"
    workbook.save(sheet_name)
    return sheet_name


def mis_loan_portfolio_snapshot_report(id, filters=None, emails=[]):
    """
    to generte loan portfolio snapshot
    :return:
    """
    disbursementStatusTracking = get_data_from_workflow(
        "disbursement_status_tracking", filters)
    if disbursementStatusTracking.get('status') == False or not disbursementStatusTracking.get('status'):
        return disbursementStatusTracking
    disbursementStatusTracking = disbursementStatusTracking.get('response')
    customerJourneyTracking = get_data_from_workflow(
        "customer_journey_tracking", filters)
    if customerJourneyTracking.get('status') == False or not customerJourneyTracking.get('status'):
        return customerJourneyTracking
    customerJourneyTracking = customerJourneyTracking.get('response')
    workbook = openpyxl.load_workbook(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'doc-template/loan_portfolio.xlsx'))
    work_sheet = workbook['Loan Portfolio Snapshot']

    work_sheet[chr(68) + str(19)].value = filters.get("from",
                                                      "") if filters else ""
    work_sheet[chr(69) + str(19)].value = filters.get("to",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(20)].value = filters.get("product",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(21)
               ].value = filters.get("customertype", "") if filters else ""
    work_sheet[chr(68) + str(22)].value = filters.get("province",
                                                      "") if filters else ""
    work_sheet[chr(68) + str(23)].value = filters.get("platform",
                                                      "") if filters else ""

    row = 25
    for email in emails:
        work_sheet[chr(66) + str(row)].value = email
        row += 1

    work_sheet[chr(70) + str(7)].value = format_number(
        disbursementStatusTracking['data']['approvedLoans']['total']['number'])
    work_sheet[chr(71) + str(7)].value = format_number(
        disbursementStatusTracking['data']['approvedLoans']['total']['amount'])
    work_sheet[chr(70) + str(8)].value = format_number(
        disbursementStatusTracking['data']['successfulDisbursement']['total']['number'])
    work_sheet[chr(71) + str(8)].value = format_number(
        disbursementStatusTracking['data']['successfulDisbursement']['total']['amount'])
    work_sheet[chr(70) + str(9)].value = format_number(
        disbursementStatusTracking['data']['closedDisbursement']['total']['number'])
    work_sheet[chr(71) + str(9)].value = format_number(
        disbursementStatusTracking['data']['closedDisbursement']['total']['amount'])
    work_sheet[chr(70) + str(10)].value = format_number(
        disbursementStatusTracking['data']['approvedLoans']['total']['number'] +
        disbursementStatusTracking['data']['successfulDisbursement']['total']['number'] -
        disbursementStatusTracking['data']['closedDisbursement']['total']['number'])
    # work_sheet[chr(70) + str(10)].value =
    work_sheet[chr(71) + str(10)].value = format_number(
        disbursementStatusTracking['data']['approvedLoans']['total']['amount'] +
        disbursementStatusTracking['data']['successfulDisbursement']['total']['amount'] -
        disbursementStatusTracking['data']['closedDisbursement']['total']['amount'])
    work_sheet[chr(70) + str(11)].value = format_number(
        disbursementStatusTracking['data']['approvedLoans']['total']['number'] * 100 /
        max(1, customerJourneyTracking['data']['totalLeads']['total']['number']), 1)
    work_sheet[chr(71) + str(11)].value = format_number(
        disbursementStatusTracking['data']['approvedLoans']['total']['amount'] * 100 /
        max(1, customerJourneyTracking['data']['totalLeads']['total']['amount']), 1)
    work_sheet[chr(70) + str(12)].value = format_number(
        disbursementStatusTracking['data']['successfulDisbursement']['total']['number'] * 100 /
        max(1, customerJourneyTracking['data']['totalLeads']['total']['number']), 1)
    work_sheet[chr(71) + str(12)].value = format_number(
        disbursementStatusTracking['data']['successfulDisbursement']['total']['amount'] * 100 /
        max(1, customerJourneyTracking['data']['totalLeads']['total']['amount']), 1)
    work_sheet[chr(70) + str(13)].value = format_number(
        disbursementStatusTracking['data']['closedDisbursement']['total']['number'] * 100 /
        max(1, disbursementStatusTracking['data']['approvedLoans']['total']['number']), 1)
    work_sheet[chr(71) + str(13)].value = format_number(
        disbursementStatusTracking['data']['closedDisbursement']['total']['amount'] * 100 /
        max(1, disbursementStatusTracking['data']['approvedLoans']['total']['amount']), 1)

    os.makedirs(MIS_SHARED_PATH, exist_ok=True)
    sheet_name = MIS_SHARED_PATH + str(id).format("loan_portfolio") + ".xlsx"
    workbook.save(sheet_name)
    return sheet_name


# def send_mail(sheet_name, to_emails):
#     if sheet_name == None or sheet_name == "" or len(to_emails) == 0:
#         return {"error": "error"}
#     file = open(sheet_name, 'rb')
#     to_emails = to_emails
#     email_params = {
#         "request_id": uuid.uuid4(),
#         "to_emails": [to_emails],
#         "from_email": "support_robo@fecredit.com.vn",
#         "subject": "MIS Report FE Credit",
#         "message": "Please find the attachment below.",
#         "attachments": [{
#             'file_name': sheet_name,
#             'file_content': base64.b64encode(file.read())
#         }]
#     }
#     url = os.environ.get('FE_CREDIT_INTEGRATION_BROKER_IP_ADDRESS',
#                          'http://localhost:8000') + "/integrations/api/email/smtp/send/"
#     response = requests.post(url, json=email_params)
#     return response.json()

def get_payment_mode(obj):
    """
    utility to get payment mode
    :param obj:
    :return:
    """
    ref_num_first_two_characters = obj.REF_NUM[:2]
    try:
        int(ref_num_first_two_characters)
    except ValueError:
        channel = fec_dbsync_cons.CHANNEL_DICT.get(
            ref_num_first_two_characters, '')
    else:
        channel = 'VNPOST'
    return '{0} {1}'.format(channel, obj.REF_NUM)


def get_transaction_history(loan_acct_num):
    dmd_dates = RoboIdtView.objects.filter(LOAN_ACCT_NUM=loan_acct_num).values('DMD_EFF_DATE'). \
        annotate(adjustementDate=Max('LAST_ADJ_DATE')).order_by('DMD_EFF_DATE')
    repay_objects = RoboRepayUplView.objects.filter(
        LOAN_ACCT_NUM=loan_acct_num).order_by('-PAID_DATE')
    repay_object_list = []
    last_adj_date = None
    for repay_object in repay_objects:
        flag = 1
        for i in range(1, len(dmd_dates)):
            if dmd_dates[i - 1]['DMD_EFF_DATE'] < repay_object.PAID_DATE <= dmd_dates[i]['DMD_EFF_DATE']:
                flag = 0
                repay_object_list.append(
                    {"repay_object": repay_object, "last_adj_date": dmd_dates[i]['adjustementDate']})
                break
        if flag:
            if not last_adj_date:
                last_adj_date = RoboIdtView.objects.filter(LOAN_ACCT_NUM=loan_acct_num).values(
                    'LAST_ADJ_DATE').aggregate(Min('LAST_ADJ_DATE'))
            repay_object_list.append(
                {"repay_object": repay_object, "last_adj_date": last_adj_date['LAST_ADJ_DATE__min']})
    return repay_object_list


def get_overdue_amount(robo_idt_view):
    overdue_amount = 0
    for idt in robo_idt_view:
        if idt['DMD_AMT_SUM'] - idt['TOT_ADJ_AMT_SUM'] > 0:
            overdue_amount = idt['DMD_AMT_SUM'] - idt['TOT_ADJ_AMT_SUM']
    return overdue_amount

# nguyenhuuvinh4 START


def get_data_from_cust_con_add(contract_number):
    SQL = "SELECT s.CONTRACT_NO,s.FULL_ADDRESS,s.CCDM_CUST_ID,s.CUST_NAME,s.TYPE_OF_NO,s.UNQ_ID,s.DISBURSAL_DT FROM ccdm.CUST_CON_ADD s WHERE s.CONTRACT_NO = '{}'"
    SQL = SQL.format(contract_number)
    print(SQL)
    connection = cx_Oracle.connect(
        "PATTA_RAJESH", "Pa55w0rd_$U", "172.27.1.7:1521/dwproddc", encoding="utf-8", nencoding="UTF-8")
    cursor = connection.cursor()
    cursor.execute(SQL)
    row = cursor.fetchone()
    response = {}
    count = 0
    for col in cursor.description:
        response[col[0]] = row[count]
        count = count+1
    cursor.close()
    connection.close()
    return response
# nguyenhuuvinh4 END


def getActiveLoanByAppIds(appIds):
    SQL = "SELECT LOAN_ACCT_NUM,clr_bal_amt as REMAINING_BAL,a.LOAN_AMOUNT,'A' as STATUS,'PL' as PRODUCT_CODE FROM custom.LOANGEN_DTLS_TABLE@ods_finnacleprod a left join tbaadm.gam@ods_finnacleprod b on a.LOAN_ACCT_NUM = b.FORACID where b.clr_bal_amt <> 0 and a.agreement_id in ({})"
    SQL = SQL.format(appIds)
    print(SQL)
    connection = cx_Oracle.connect(
        "PATTA_RAJESH", "Pa55w0rd_$U", "172.27.1.7:1521/dwproddc", encoding="utf-8", nencoding="UTF-8")
    cursor = connection.cursor()
    cursor.execute(SQL)
    response = []
    for row in cursor:
        response.append(row)
    cursor.close()
    connection.close()
    return response


def getFraudDataFromOracle(fraudRuleName, value, zoneRangeFlag, productCode):

    SQL = ""
    mydb = mysql.connector.connect(host="192.169.219.82",
                                   user="root",
                                   passwd="Robo@123",
                                   auth_plugin='mysql_native_password')
    data = None
    cursor = mydb.cursor()
    if not zoneRangeFlag:
        if not productCode:
            SQL = "select ACTION from Workflow.fraudData where PARAMETER_KEY='{}' and VALUE='{}' and CURRENT_TIMESTAMP between ACTIONFROM  and ACTIONTO"
        else: 
            SQL = "select ACTION from Workflow.fraudData where PARAMETER_KEY='{}' and VALUE='{}' and PRODUCTID ='{}' and CURRENT_TIMESTAMP between ACTIONFROM  and ACTIONTO"
        SQL = SQL.format(fraudRuleName, value, productCode)
        print("Query : ", SQL)
        cursor.execute(SQL)
        listData = list(cursor.fetchall())
        data = []
        for li in listData:
            data.append(li[0])
    else:
        SQL = "select VALUE, ACTION from Workflow.fraudData where PARAMETER_KEY='locationZone' and CURRENT_TIMESTAMP between ACTIONFROM  and ACTIONTO"
        cursor.execute(SQL)
        listData = list(cursor.fetchall())
        dicVal = {}
        data = {}
        for li in listData:
            data[li[0]] = li[1]
    #print("Fraud rule : ",fraudRuleName+" value : "+value )
    print('Data found ---:', data)
    cursor.close()
    mydb.close()
    return data


def getFraudDataFromMysql(fraudRuleName, value, zoneRangeFlag):

    SQL = ""
    mydb = mysql.connector.connect(host="192.169.219.82",
                                   user="root",
                                   passwd="Robo@123",
                                   auth_plugin='mysql_native_password')
    data = None
    cursor = mydb.cursor()
    if not zoneRangeFlag:
        SQL = "select ACTION from Workflow.fraudData where PARAMETER_KEY='{}' and VALUE='{}' and CURRENT_TIMESTAMP between ACTIONFROM  and ACTIONTO"
        SQL = SQL.format(fraudRuleName, value)
        cursor.execute(SQL)
        listData = list(cursor.fetchall())
        data = []
        for li in listData:
            data.append(li[0])
    else:
        SQL = "select VALUE, ACTION from Workflow.fraudData where PARAMETER_KEY='locationZone' and CURRENT_TIMESTAMP between ACTIONFROM  and ACTIONTO"
        cursor.execute(SQL)
        listData = list(cursor.fetchall())
        dicVal = {}
        data = {}
        for li in listData:
            data[li[0]] = li[1]
    print(data)
    cursor.close()
    mydb.close()
    return data


def getBlackListedBankAccountsFromOracle(bank_account):
    try:
        connection = cx_Oracle.connect(
            "PATTA_RAJESH", "Pa55w0rd_$U", "172.27.1.7:1521/dwproddc", encoding="utf-8", nencoding="UTF-8")
        cursor = connection.cursor()

        sql = """select
                "TASK_ID",
                "BANK_ACCOUNT",
                "CUST_NAME",
                "EFFECT_FROM",
                "EFFECT_TO",
                "STATUS",
                "CASE_NO"
            from Operation_risk.af_tbl_blacklist_bank_account
            where
                "BANK_ACCOUNT" = :bank_account and
                "STATUS" = '1'
            """

        cursor.execute(sql, bank_account = bank_account)

        print("cursor = ", cursor)

        data = []

        for (TASK_ID, BANK_ACCOUNT, CUST_NAME, EFFECT_FROM, EFFECT_TO, STATUS, CASE_NO) in cursor:
            print("TASK_ID = ", cursor)

            item = {
                "task_id": TASK_ID,
                "bank_account": BANK_ACCOUNT,
                "cust_name": CUST_NAME,
                "effect_from": EFFECT_FROM,
                "effect_to": EFFECT_TO,
                "status": STATUS,
                "case_no": CASE_NO
            }

            data.append(item)

        print('getBlackListedBankAccountFromOracle > Data: ', data)

        cursor.close()
        connection.close()

        return data
    except Exception as e:
        print("Exception: ", e)

        return []

def getBlackListedBankAccountsFromMySQL(bankAccount):
    connection = mysql.connector.connect(host="192.169.219.82",
                                         user="root",
                                         passwd="Robo@123",
                                         auth_plugin='mysql_native_password')
    cursor = connection.cursor()

    sql = """select
            TASK_ID,
            BANK_ACCOUNT,
            CUST_NAME,
            EFFECT_FROM,
            EFFECT_TO,
            STATUS,
            CASE_NO
        from `fec_db`.`fec_dbsync_af_tbl_blacklist_bank_account`
        where
            BANK_ACCOUNT = %(bank_account)s and
            STATUS = '1'
        """

    cursor.execute(sql, {"bank_account": bankAccount})

    data = []

    for (TASK_ID, BANK_ACCOUNT, CUST_NAME, EFFECT_FROM, EFFECT_TO, STATUS, CASE_NO) in cursor:
        item = {
            "task_id": TASK_ID,
            "bank_account": BANK_ACCOUNT,
            "cust_name": CUST_NAME,
            "effect_from": EFFECT_FROM,
            "effect_to": EFFECT_TO,
            "status": STATUS,
            "case_no": CASE_NO
        }

        data.append(item)

    print('getBlackListedBankAccountsFromMySQL > Data: ', data)

    cursor.close()
    connection.close()

    return data


def calculateDistance(lat1, lat2, long1, long2):
    R = 6373.0
    lat1 = radians(float(lat1))
    lat2 = radians(float(lat2))
    lon1 = radians(float(long1))
    lon2 = radians(float(long2))

    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))
    distance = R * c * 1000
    print("distance in meter : ", distance)
    return distance


def check(data, req, dataKey):
    keysFlags = req['dataFlags']
    keyFraud = False

    fraudF = req['fraudFlag']
    softRejectF = req['softRejectFlag']
    freezeF = req['freezeFlag']
    dataFraudF = req['dataFraudFlag']
    if data != None and len(data) > 0:
        keyFraud = dataFraudF = True
        if not softRejectF and 'Soft Reject' in data:
            softRejectF = True
        elif not fraudF and 'Fraud' in data:
            fraudF = True
        elif not freezeF and 'Freeze' in data:
            freezeF = True
    response = {}
    keysFlags[dataKey] = keyFraud
    response.update(dataFlags=keysFlags)
    response.update(fraudFlag=fraudF, softRejectFlag=softRejectF,
                    freezeFlag=freezeF, dataFraudFlag=dataFraudF)
    return response

def checkAMLService(req):

    # UAT
    SQL_Order1 = "SELECT * FROM fec_db.fec_dbsync_amltable WHERE PERSON_ID = '{}' AND STATUS = 1"
    SQL_Order2 = "SELECT * FROM fec_db.fec_dbsync_amltable WHERE FULL_NAME = '{}' AND ID_NO = '{}' AND STATUS = 1"
    SQL_Order3 = "SELECT * FROM fec_db.fec_dbsync_amltable WHERE FULL_NAME = '{}' AND DOB = '{}' AND STATUS = 1"
    SQL_Order4 = "SELECT * FROM fec_db.fec_dbsync_amltable WHERE FULL_NAME = '{}' AND STATUS = 1"
    # print(SQL)
    SQL = ""
    isReject = True
    response = {}
    print("isReject = ", isReject)

    if req.data.get('PERSON_ID') != None:
        # Check person_ID all applications
        SQL  = SQL_Order1.format(req.data.get('PERSON_ID'))
        print (SQL)
        isReject = OracleService(SQL)

        if not isReject:
            # checked against the combination of Name and National ID
            SQL = SQL_Order2.format(req.data.get('FULL_NAME'), req.data.get('ID_NO'))
            print (SQL)
            isReject = OracleService(SQL)

            if not isReject:
                # check against the combination of Name and date of birth
                SQL = SQL_Order3.format(req.data.get('FULL_NAME'), req.data.get('DOB'))
                print (SQL)
                isReject = OracleService(SQL)

                if not isReject:
                    #  checked against these Name in AML table
                    SQL = SQL_Order4.format(req.data.get('FULL_NAME'))
                    print (SQL)
                    isReject = OracleService(SQL)

                    if isReject:
                        isReject = False
                        response.update(amlFlag=True)

    response.update(data=SQL,isReject=isReject)
    return response

def blacklistPhoneNumber(req):
    SQL_1 = "SELECT * FROM fec_db.af_tbl_blacklist_phone WHERE CODE_NUMBER = '{}' AND STATUS = 1 AND (EFFECT_TO >= SYSDate() OR EFFECT_TO IS NULL);"
    isReject = True
    SQL = ""
    response = {}

    if req.data.get('phoneNumber') != None:
        SQL  = SQL_1.format(req.data.get('phoneNumber'))
        isReject  = OracleService(SQL)

    response.update(data=SQL,isReject=isReject)
    return response

def OracleService(sql):
    #UAT
    connection = mysql.connector.connect(host="192.169.219.82",
                                   user="root",
                                   passwd="Robo@123",
                                   auth_plugin='mysql_native_password')
    cursor = connection.cursor()
    cursor.execute(sql)
    isReject = False
    data = []

    for row in cursor:
        data.append(row)

    cursor.close()
    connection.close()

    if data:
        print(data)
        isReject = True

    return isReject
      