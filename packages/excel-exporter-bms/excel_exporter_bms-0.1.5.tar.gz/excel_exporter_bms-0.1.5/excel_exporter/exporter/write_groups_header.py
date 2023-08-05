from typing import Dict, Tuple

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from excel_exporter.configuration.group import Group as GroupConfiguration


def write_groups_header(
    ws: Worksheet,
    groups_config: Dict[str, GroupConfiguration],
    groups_limits: Dict[str, Tuple],
    start_col: int = 2,
    row_number: int = 2,
):
    alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    font = Font(bold=True, color='FFFFFF')
    for group_name, group_limits in groups_limits.items():
        group = groups_config[group_name]
        color = group.background_color.lstrip('#')
        fill = PatternFill(
            start_color=color, end_color=color, fill_type='solid'
        )
        group_start_col = group_limits[0] + start_col - 1
        group_end_col = group_limits[1] + start_col - 1
        ws.merge_cells(
            start_row=row_number,
            end_row=row_number,
            start_column=group_start_col,
            end_column=group_end_col,
        )
        cell = ws.cell(row=row_number, column=group_start_col)
        cell.value = group.group_name
        cell.fill = fill
        cell.alignment = alignment
        cell.font = font
