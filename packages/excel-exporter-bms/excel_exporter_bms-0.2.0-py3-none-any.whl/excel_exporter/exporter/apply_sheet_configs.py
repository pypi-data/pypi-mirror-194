import re

from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from excel_exporter.configuration.sheet import Sheet as SheetConfiguration


def make_table_name(sheet_name: str) -> str:
    table_name = sheet_name.replace(' ', '_').lower()
    table_name = re.sub(r'[^a-zA-Z0-9_]+', '', table_name)
    return table_name


def apply_sheet_configs(ws: Worksheet, ws_config: SheetConfiguration):
    # Sheet name
    ws.title = ws_config.sheet_name
    # Format Table
    ref = f'B3:{get_column_letter(ws.max_column)}{str(ws.max_row)}'
    table_name = make_table_name(ws_config.sheet_name)
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name='TableStyleMedium1',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    # Adjust border columns widths
    left_border_col = get_column_letter(ws.min_column - 1)
    right_border_col = get_column_letter(ws.max_column + 1)
    ws.column_dimensions[left_border_col].width = 2
    ws.column_dimensions[right_border_col].width = 2
    # Hide gridlines
    ws.sheet_view.showGridLines = False
    # Hide unused columns:
    hidden_begin = ws.max_column + 2
    hidden_end = column_index_from_string('XFD') + 1
    for col_number in range(hidden_begin, hidden_end):
        col_letter = get_column_letter(col_number)
        ws.column_dimensions[col_letter].hidden = True
