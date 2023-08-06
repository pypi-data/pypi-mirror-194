import openpyxl
from openpyxl.styles import Alignment

import math
import pandas as pd
import sys

def default():
    return openpyxl.Workbook()

def from_file(file):
    return openpyxl.load_workbook(file)

def from_DataFrames(dataFrames):
    dataFrames = dict(dataFrames)
    if len(dataFrames) == 0:
        return None
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    for table, df in dataFrames.items():
        if default_sheet is None:
            workbook.create_sheet(table)
        else:
            default_sheet.title = table
            default_sheet = None
    for table, df in dataFrames.items():
        columns = list(df.columns)
        for x, column in enumerate(columns):
            workbook[table].cell(row=1, column=x+1).value = column
            for y, v in enumerate(df[column].tolist()):
                if pd.isna(v):
                    continue
                elif (type(v) is float) and (math.isinf(v)):# is this really needed?
                    value = str(v)
                else:
                    value = v
                workbook[table].cell(row=y+2, column=x+1).value = value
    return workbook

def set_cell(*, cell, value):
    """Setting value of cell. """
    if pd.isna(value):
        value = 'N/A'
    else:
        if type(value) is float:
            if math.isinf(value):
                if value < 0:
                    value = '-inf'
                else:
                    value = '+inf'
        if type(value) not in {str, int, float, bool}:
            raise TypeError(f"The value {value} is of the invalid type {type(value)}! ")
    cell.value = value
    cell.alignment = Alignment()#horizontal='general')


