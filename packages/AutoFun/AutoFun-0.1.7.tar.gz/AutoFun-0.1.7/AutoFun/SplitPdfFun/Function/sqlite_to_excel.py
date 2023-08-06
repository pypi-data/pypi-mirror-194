try:
    from .DB_in_sqlite import get_df_from_sqlite_OCRtxt
except:
    from DB_in_sqlite import get_df_from_sqlite_OCRtxt
from openpyxl import load_workbook
import re
import pandas as pd
#所有的索引均为数字加杠加数字形式，杠可以为零，可以为第几章的形式，其他类型索引不支持。
pd.set_option('display.max_columns',None)
# 显示所有列
pd.set_option('display.max_rows',None)
# 设置数据的显示长度，默认为50
pd.set_option('max_colwidth',200)
# 禁止自动换行(设置为Flase不自动换行，True反之)
pd.set_option('expand_frame_repr', False)

LAYRER_FILE=6
AUTO_LAYER=False

#阿拉伯数字转换汉字数字的函数
def change_num_ArabToChinese(i):
    i=int(i)
    num_dict = {'1': '一', '2': '二', '3': '三', '4': '四', '5': '五', '6': '六', '7': '七', '8': '八', '9': '九', '0': '零', }
    index_dict = {1: '', 2: '十', 3: '百', 4: '千', 5: '万', 6: '十', 7: '百', 8: '千', 9: '亿'}
    if 10<=i<=19:
        i=str(i)
        nums = list(i)
        nums_index = [x for x in range(1, len(nums) + 1)][-1::-1]

        str0 = ''
        for index, item in enumerate(nums):
            str0 = "".join((str0, num_dict[item], index_dict[nums_index[index]]))

        str0 = re.sub("零[十百千零]*", "零", str0)
        str0 = re.sub("零万", "万", str0)
        str0 = re.sub("亿万", "亿零", str0)
        str0 = re.sub("零零", "零", str0)
        str0 = re.sub("零\\b", "", str0)
        str0=str0[1:]
        return str0
    else:
        i = str(i)

        nums = list(i)
        nums_index = [x for x in range(1, len(nums) + 1)][-1::-1]

        str0 = ''
        for index, item in enumerate(nums):
            str0 = "".join((str0, num_dict[item], index_dict[nums_index[index]]))

        str0 = re.sub("零[十百千零]*", "零", str0)
        str0 = re.sub("零万", "万", str0)
        str0 = re.sub("亿万", "亿零", str0)
        str0 = re.sub("零零", "零", str0)
        str0 = re.sub("零\\b", "", str0)
        return str0

#寻找是否包含不能命名的字符
def find_ilegal_character(txt):
    list_ilegal_character=[]
    if '\\' in txt:
        list_ilegal_character.append('\\')
    if '/' in txt:
        list_ilegal_character.append('/')
    if ':' in txt:
        list_ilegal_character.append(':')
    if '*' in txt:
        list_ilegal_character.append('*')
    if '?' in txt:
        list_ilegal_character.append('?')
    if '"' in txt:
        list_ilegal_character.append('"')
    if '<' in txt:
        list_ilegal_character.append('<')
    if '>' in txt:
        list_ilegal_character.append('>')
    if '|' in txt:
        list_ilegal_character.append('|')
    if '\n' in txt:
        list_ilegal_character.append('回车')
    if '\t' in txt:
        list_ilegal_character.append('tab')
    return list_ilegal_character

#从sqlite导出excel
def gen_ori_ocr_excel():
    df_ocr=get_df_from_sqlite_OCRtxt()
    df_ocr.to_excel('ori_tb_OCR.xlsx',index=False)

#从文本中抽取出几杠几和第几章的索引结构
def find_index(ocrtxt):
    if re.match('(\d+(?:-\d+)*)',ocrtxt):
        # ws.cell(row=i, column=6).value='是'
        isindex=True
        ocr_index=re.match('(\d+(?:-\d+)*)',ocrtxt).group(1)
        # ws.cell(row=i, column=7).value = ocr_index
        if ocr_index.count('-')==0:
            if int(ocr_index)<99:
                ocr_index = re.match('(\d+(?:-\d+)*)', ocrtxt).group(1)
                n_layers = ocr_index.count('-') + 1
            else:
                ocr_index = ''
                n_layers = ''
        else:
            ocr_index = re.match('(\d+(?:-\d+)*)', ocrtxt).group(1)
            n_layers = ocr_index.count('-')+1
        # ws.cell(row=i, column=8).value = str(n_line+1)
    elif re.match('(第[一二三四五六七八九十1234567890 ]{1,4}章)',ocrtxt):
        ocr_index = re.match('(第[一二三四五六七八九十1234567890 ]{1,5}章)', ocrtxt).group(1).replace(' ','')
        n_layers = 1
        isindex = True
    else:
        isindex = False
        ocr_index=''
        n_layers=''
    index_list=[isindex,ocr_index,n_layers]
    return index_list

#格式化调整索引和名称格式
def regularize_index_name(ocrtxt):
    if not ocrtxt:
        ocrtxt=''
    if re.match('(\d+(?:-\d+)*)', ocrtxt):
        # ws.cell(row=i, column=6).value='是'
        isindex = True
        ocr_index = re.match('(\d+(?:-\d+)*)', ocrtxt).group(1)
        ocr_name=re.match('\d+(?:-\d+)*(.*)', ocrtxt).group(1).strip()
        # ws.cell(row=i, column=7).value = ocr_index
        if ocr_index.count('-') == 0:
            if int(ocr_index) < 99:
                ocr_index = re.match('(\d+(?:-\d+)*)', ocrtxt).group(1)
                ocr_name = re.match('\d+(?:-\d+)*(.*)', ocrtxt).group(1).strip()
                n_layers = ocr_index.count('-') + 1
            else:
                ocr_index = ''
                ocr_name = ocrtxt
                n_layers = ''
        else:
            ocr_index = re.match('(\d+(?:-\d+)*)', ocrtxt).group(1)
            ocr_name = re.match('\d+(?:-\d+)*(.*)', ocrtxt).group(1).strip()
            n_layers = ocr_index.count('-') + 1
        # ws.cell(row=i, column=8).value = str(n_line+1)
    elif re.match('(第[一二三四五六七八九十1234567890 ]{1,4}章)', ocrtxt):
        ocr_index = re.match('(第[一二三四五六七八九十1234567890 ]{1,5}章)', ocrtxt).group(1).replace(' ', '')
        ocr_name = re.match('第[一二三四五六七八九十1234567890 ]{1,5}章(.*)', ocrtxt).group(1).strip()
        n_layers = 1
        isindex = True
    else:
        isindex = False
        ocr_index = ''
        ocr_name = ocrtxt
        n_layers = ''
    regularize_index_name=(ocr_index+'  '+ocr_name).strip()
    return regularize_index_name


#在excel里填充表头，寻找索引
def fill_tb_index(tb_path):
    wb=load_workbook(tb_path)
    ws=wb['Sheet1']
    ws.cell(row=1, column=1).value ='文件名'
    ws.cell(row=1, column=2).value ='文本内容'
    ws.cell(row=1, column=3).value ='页码'
    ws.cell(row=1, column=4).value ='下限文本置信度'
    ws.cell(row=1, column=5).value ='平均文本置信度'
    ws.cell(row=1, column=6).value ='是否有索引结构'
    ws.cell(row=1, column=7).value ='索引内容'
    ws.cell(row=1, column=8).value ='索引层级'
    ws.cell(row=1, column=9).value ='上一级别索引'
    ws.cell(row=1, column=10).value ='命名合法性检测'
    ws.cell(row=1, column=11).value = '切分后文件名'
    i=2
    num_null=0
    filename_pre=''
    while True:
        if num_null>50:
            break
        if not ws.cell(row=i, column=1).value:
            num_null+=1
            i += 1
            continue
        filename_now = ws.cell(row=i, column=1).value
        if filename_now!=filename_pre:
            filename_pre=filename_now
            if int(ws.cell(row=i, column=3).value)!=1:
                ws.insert_rows(i)
                ws.cell(row=i, column=1).value = ws.cell(row=i+1, column=1).value
                ws.cell(row=i, column=3).value =str(1)
                i+=1
                continue
        ocrtxt=str(ws.cell(row=i, column=2).value)
        list_index=find_index(ocrtxt)
        if list_index[0]:
            ws.cell(row=i, column=6).value = '是'
            ws.cell(row=i, column=7).value = list_index[1]
            ws.cell(row=i, column=8).value = list_index[2]
        else:
            ws.cell(row=i, column=6).value = '否'
            ws.cell(row=i, column=7).value = list_index[1]
            ws.cell(row=i, column=8).value = list_index[2]
        i+=1
    # print(find_index('现金收付交易'))
    wb.save(tb_path)

#父级索引的确认维度有索引结构和页数两个维度，确认逻辑为先从抽取索引后的表中读取每个文件内的所有索引，1、索引为无，则parent为None；2、索引为整数或者第几章，则parent为None；3、索引为几杠几，则去除尾部最后一个杠数字作为父索引检索，如无则迭代，直到找到或者空，若有多个，则页数必须小于当前索引页数结果中离当前索引最近的那个是父索引。
#上面的思路有漏洞 比如一个文件里有两个树形结构 第二个树的某节点由于某种原因缺失，导致第二个文件的子节点会错误的挂到第一个树的子节点上，应该在文件处理开始之前，先判断是否有多棵树，如果有，找到分界线，在分界内处理
#上面这个理论可行，实际情况复杂多变，难以实现
def find_parent_index(tb_path):
    df_no_parent=pd.read_excel(tb_path, dtype=str)
    file_list=list(set(df_no_parent['文件名'].tolist()))
    print(file_list)
    for eachfile in file_list:
        df_each_file=df_no_parent.loc[df_no_parent['文件名']==eachfile]
        # print(df_each_file)
        list_this_file_pdindex=df_each_file.index.tolist()
        # print(list_this_file_pdindex)
        parent_index=None
        for each_pd_index in list_this_file_pdindex:
            # print(df_no_parent.loc[each_pd_index])
            #在这里判断几棵树，并且把树分开，然后遍历每棵树
            #上面的想法其实在实用层面上，可能是概念模糊的，暂时放弃，
            #
            now_index=df_no_parent.loc[each_pd_index,'索引内容']
            if now_index:
                while True:
                    if str(now_index).count('-')==0:
                        parent_index = None
                        break
                    elif str(now_index).count('-')==1:
                        parent_index=re.match('((\d+(?:-\d+)*))-\d+', now_index).group(1)
                        df_par = df_each_file.loc[df_each_file['索引内容'] == parent_index]
                        if df_par.shape[0]:
                            if df_par.shape[0]>1:
                                df_par=df_par.loc[df_par['页码']<df_no_parent.loc[each_pd_index,'页码']]
                                if df_par.shape[0]:
                                    df_par=df_par.sort_values(by=['页码'],ascending=False)
                                    parent_index = df_par.iloc[0, 6]
                                else:
                                    df_par = df_par.sort_values(by=['页码'], ascending=True)
                                    parent_index = df_par.iloc[0, 6]
                                break
                            else:
                                parent_index =df_par.iloc[0,6]
                                break
                        else:
                            parent_txt2='第'+change_num_ArabToChinese(parent_index) +'章'
                            df_par = df_each_file.loc[df_each_file['索引内容'] == parent_txt2]
                            if df_par.shape[0]:
                                if df_par.shape[0] > 1:
                                    df_par = df_par.loc[df_par['页码'] < df_no_parent.loc[each_pd_index, '页码']]
                                    if df_par.shape[0]:
                                        df_par = df_par.sort_values(by=['页码'], ascending=False)
                                        parent_index = df_par.iloc[0, 6]
                                    else:
                                        df_par = df_par.sort_values(by=['页码'], ascending=True)
                                        parent_index = df_par.iloc[0, 6]
                                    break
                                else:
                                    parent_index = df_par.iloc[0, 6]
                                    break
                            else:
                                parent_txt1 = '第' + parent_index + '章'
                                df_par = df_each_file.loc[df_each_file['索引内容'] == parent_txt1]
                                if df_par.shape[0]:
                                    if df_par.shape[0] > 1:
                                        df_par = df_par.loc[df_par['页码'] < df_no_parent.loc[each_pd_index, '页码']]
                                        if df_par.shape[0]:
                                            df_par = df_par.sort_values(by=['页码'], ascending=False)
                                            parent_index = df_par.iloc[0, 6]
                                        else:
                                            df_par = df_par.sort_values(by=['页码'], ascending=True)
                                            parent_index = df_par.iloc[0, 6]
                                        break
                                    else:
                                        parent_index = df_par.iloc[0, 6]
                                        break
                                else:
                                    parent_index = None
                                    break
                    else:
                        parent_index=re.match('((\d+(?:-\d+)*))-\d+', now_index).group(1)
                        df_par=df_each_file.loc[df_each_file['索引内容']==parent_index]
                        if df_par.shape[0]:
                            if df_par.shape[0] > 1:
                                df_par = df_par.loc[df_par['页码'] < df_no_parent.loc[each_pd_index, '页码']]
                                if df_par.shape[0]:
                                    df_par = df_par.sort_values(by=['页码'], ascending=False)
                                    parent_index = df_par.iloc[0, 6]
                                else:
                                    df_par = df_par.sort_values(by=['页码'], ascending=True)
                                    parent_index = df_par.iloc[0, 6]
                                break
                            else:
                                parent_index = df_par.iloc[0, 6]
                                break
                        else:
                            now_index=parent_index
            else:
                parent_index=None
            df_no_parent.loc[each_pd_index,'上一级别索引']=parent_index
            txt_each_pd_index=str(df_no_parent.loc[each_pd_index,'文本内容'])
            list_ilegal=find_ilegal_character(txt_each_pd_index)
            txt_read=''
            if len(list_ilegal):
                txt_read='检测到包含不可命名字符：【'
                for i_each_ilegal in range(0,len(list_ilegal)):
                    txt_read+=list_ilegal[i_each_ilegal]
                    if i_each_ilegal<len(list_ilegal)-1:
                        txt_read +='、'
                    else:
                        txt_read += '】。\n'
            if len(txt_each_pd_index)>69:
                txt_read += '文本长度大于等于70，建议精简。'
            df_no_parent.loc[each_pd_index, '命名合法性检测'] = txt_read
            # print(parent_index)
    # print(df_no_parent)
    # df_no_parent.to_excel('ori_tb_OCR_with_parent.xlsx',index=False)
    df_no_parent.to_excel(tb_path,index=False)

# def compute_new_filename_new_loc_pd(tb_path,layer_file=LAYRER_FILE,auto_layer=AUTO_LAYER):
#     df_old_file = pd.read_excel(tb_path, dtype=str)
#     file_list = list(set(df_old_file['文件名'].tolist()))
#     print(file_list)
#     for eachfile in file_list:
#         df_each_file = df_old_file.loc[df_old_file['文件名'] == eachfile]
#         # print(df_each_file)
#         list_this_file_pdindex = df_each_file.index.tolist()
#         # print(list_this_file_pdindex)
#         parent_index = None
#         for each_pd_index in list_this_file_pdindex:
#             # print(df_no_parent.loc[each_pd_index])
#             # 在这里判断几棵树，并且把树分开，然后遍历每棵树
#             # 上面的想法其实在实用层面上，可能是概念模糊的，暂时放弃，
#             #
#             now_index = df_old_file.loc[each_pd_index, '索引内容']
#             now_n_page= df_old_file.loc[each_pd_index, '页码']
#             now_txt=df_old_file.loc[each_pd_index, '文本内容']
#             new_filename='[WARNING]'+str(df_old_file.loc[each_pd_index, '文件名'])+'自'+str(now_n_page)+'页起始的文件'
#             if not now_index and int(now_n_page)==1:
#                 df_old_file.loc[each_pd_index, '切分后文件名']=new_filename
#                 continue
#             if now_index:

def compute_new_filename_new_loc(tb_path, layer_file=LAYRER_FILE, auto_layer=AUTO_LAYER):
    wb = load_workbook(tb_path)
    ws = wb['Sheet1']
    i = 2
    num_null = 0
    filename_pre = ''
    startlinefile =0
    endlinefile =0
    sign_continue_read_line=1
    cursor_read=i
    ori_layer_file=layer_file
    while True:
        #逐行读，确定区间
        filename_now = ws.cell(row=i, column=1).value
        if i==2:
            filename_pre=filename_now

        if sign_continue_read_line:
            # i=cursor_read

            if i == cursor_read:
                startlinefile = i
            if filename_now != filename_pre:
                filename_pre=filename_now
                endlinefile=i-1
                cursor_read = i
                sign_continue_read_line = 0
                continue
            if num_null > 50:
                break#读取五十个空行后自动结束
            if not ws.cell(row=i, column=1).value:
                num_null += 1
                i += 1
                continue
            i+=1
        else:
            #在找寻好的区间内处理文件
            print('--------------------------')
            print(startlinefile)
            print(endlinefile)
            if startlinefile==142:
                print('')

            # for j in range(startlinefile,endlinefile+1):
            j=startlinefile
            while startlinefile<=j<endlinefile+1:
                #页面不漏，因此j每次移动都一定是文件名应该命名的地方，但文件名可能会在上下浮动的位置，应该是向下浮动
                now_index=ws.cell(row=j, column=7).value
                now_n_page=int(ws.cell(row=j, column=3).value)
                now_src_file_name=ws.cell(row=j, column=1).value
                if now_n_page==1:
                    if now_index:
                        #分三种情况
                        layers_now_index = int(ws.cell(row=j, column=8).value)
                        if layers_now_index>layer_file:
                            new_filename = '[WARNING]' + str(now_src_file_name) + '自' + str(now_n_page) + '页起始的过深级别索引文件，过深级别索引为'+str(now_index)
                            ws.cell(row=j, column=11).value = new_filename
                            j=j+1
                            # layer_file=ori_layer_file
                            continue
                        elif layers_now_index<layer_file:
                            #首页层级比指定层级浅，需要从下一行开始找文件名，找到指定层级或页面不连续断开即获得文件名
                            # for k in range(j + 1, endlinefile + 1):
                            k=j + 1
                            #如果下一行的k超过原文件范围，则结束
                            if k>=endlinefile + 1:
                                new_filename = ws.cell(row=j, column=2).value
                                new_filename = regularize_index_name(new_filename)
                                ws.cell(row=j, column=11).value = new_filename
                                break
                            sign_k_over=0
                            #如果下一行的k不超过原文件范围，则开始搜寻文件名是啥，搜寻的k，必须和之前的隔页对应的页码是连续的，获取文件名的条件，k对应索引足够深，或者k对应索引不连续，用k_pre和k_next来判断是否级别深度
                            while j + 1<=k<endlinefile + 1:
                                k_pre=k-1
                                k_next=k
                                #判断级别深度的前提是有索引结构，这种情况前提是j页层级浅于指定层级，k_pre从k-1即j开始，k_pre对应一定有层级，下面的判断纯属脱裤子放屁
                                while not ws.cell(row=k_pre, column=8).value:
                                    k_pre-=1
                                #判断级别深度的前提是有索引结构，k_next应该有索引，要先找到有索引的k_next，遇到指定层级应该停下来，把前面的文件存下来并标注不同层级隔页中间夹着的文件，如果遇到同级别或更浅级别文件应该保存对应文件
                                while not ws.cell(row=k_next, column=8).value:
                                    if k_next<endlinefile + 1:
                                        k_next+=1
                                    else:
                                        new_filename=ws.cell(row=k_pre, column=2).value
                                        new_filename = regularize_index_name(new_filename)
                                        ws.cell(row=j, column=11).value = new_filename
                                        j = endlinefile + 1
                                        sign_k_over=1
                                        break
                                if sign_k_over:
                                    break
                                #走到这里代表通常情况，找到了k_pre,k_next，接下来判断层级
                                pre_n_page=int(ws.cell(row=k_pre, column=3).value)
                                next_n_page=int(ws.cell(row=k_next, column=3).value)
                                # if next_n_page==(pre_n_page+1):
                                if next_n_page==(pre_n_page+k_next-k_pre):
                                # if next_n_page==(pre_n_page+k_next-k_pre-1):
                                    #判断下一行是否更深
                                    layers_next_index=int(ws.cell(row=k_next, column=8).value)
                                    layers_pre_index = int(ws.cell(row=k_pre, column=8).value)
                                    if layers_next_index>layers_pre_index:
                                        #此时应该继续深入，如果已经到末级了，就直接获取文件名
                                        if layers_next_index==layer_file:
                                            new_filename = ws.cell(row=k_next, column=2).value
                                            new_filename=regularize_index_name(new_filename)
                                            ws.cell(row=j, column=11).value = new_filename
                                            j=k_next+1
                                            # layer_file = ori_layer_file
                                            break
                                        elif layers_next_index<layer_file:
                                            if k_next==endlinefile:
                                                new_filename = ws.cell(row=k_next, column=2).value
                                                new_filename = regularize_index_name(new_filename)
                                                ws.cell(row=j, column=11).value = new_filename
                                                j = k_next + 1
                                                # layer_file = ori_layer_file
                                                break
                                            #到了这个位置，找到的k_next深度比现在深，比指定浅
                                            k=k_next+1
                                            continue
                                        else:
                                            new_filename ='[WARNING索引跨级缺失]' +regularize_index_name(ws.cell(row=k_next, column=2).value)
                                            ws.cell(row=j, column=11).value = new_filename
                                            j = k_next + 1
                                            # layer_file = ori_layer_file
                                            break
                                    #这种情况下一个的级别不深于上一个，保存文件
                                    else:
                                        new_filename=ws.cell(row=k_pre, column=2).value
                                        new_filename = regularize_index_name(new_filename)
                                        ws.cell(row=j, column=11).value = new_filename
                                        j = k_next
                                        # layer_file = ori_layer_file
                                        break
                                else:
                                    #不连续的情况是有正常情况的，当kpre是文件名时，大概率不连续
                                    layers_next_index = int(ws.cell(row=k_next, column=8).value)
                                    layers_pre_index = int(ws.cell(row=k_pre, column=8).value)
                                    if layers_next_index <= layers_pre_index:
                                        new_filename = ws.cell(row=k_pre, column=2).value
                                        new_filename = regularize_index_name(new_filename)
                                        ws.cell(row=j, column=11).value = new_filename
                                    else:
                                        new_filename = '[WARNING层级隔页之间文件]' + regularize_index_name(ws.cell(row=k_pre, column=2).value)
                                        ws.cell(row=j, column=11).value = new_filename
                                    j = k_next
                                    # layer_file = ori_layer_file
                                    break
                                    # #判断自动调节级别参数
                                    # if auto_layer:
                                    #     layer_file=layers_now_index
                                    #     # sign_temp_layer=1
                                    #     new_filename = '[WARNING索引自动调整]' + ws.cell(row=k-1, column=2).value
                                    #     j = k
                                    #     break
                                    # else:
                                    #     new_filename = '[WARNING层级隔页之间文件]' + ws.cell(row=k - 1, column=2).value
                                    #     j = k
                                    #     # layer_file = ori_layer_file
                                    #     break
                        else:
                            new_filename=ws.cell(row=j, column=2).value
                            new_filename = regularize_index_name(new_filename)
                            ws.cell(row=j, column=11).value = new_filename
                            j = j + 1
                            layer_file = ori_layer_file
                            continue
                    else:
                        new_filename = '[WARNING]' + str(now_src_file_name) + '自' + str(
                            now_n_page) + '页起始的无索引文件'
                        ws.cell(row=j, column=11).value=new_filename
                        # layer_file = ori_layer_file
                        j = j + 1
                        continue
                else:
                    #多页情况类似首页，可以参考首页处理方式
                    if now_index:
                        #分三种情况
                        layers_now_index = int(ws.cell(row=j, column=8).value)
                        if layers_now_index>layer_file:
                            j=j+1
                            layer_file=ori_layer_file
                            continue
                        elif layers_now_index<layer_file:
                            # for k in range(j + 1, endlinefile + 1):
                            #     pre_n_page=int(ws.cell(row=k-1, column=3).value)
                            #     next_n_page=int(ws.cell(row=k, column=3).value)
                            #     if next_n_page==(pre_n_page+1):
                            #         #判断下一行是否更深
                            #         layers_next_index=int(ws.cell(row=k, column=8).value)
                            #         layers_pre_index = int(ws.cell(row=k-1, column=8).value)
                            #         if layers_next_index>layers_pre_index:
                            #             #此时应该继续深入，如果已经到末级了，就直接获取文件名
                            #             if layers_next_index==layer_file:
                            #                 new_filename = ws.cell(row=k, column=2).value
                            #                 new_filename = regularize_index_name(new_filename)
                            #                 ws.cell(row=j, column=11).value = new_filename
                            #                 j=k+1
                            #                 # layer_file = ori_layer_file
                            #                 break
                            #             elif layers_next_index<layer_file:
                            #                 if k==endlinefile:
                            #                     new_filename = ws.cell(row=k, column=2).value
                            #                     new_filename = regularize_index_name(new_filename)
                            #                     ws.cell(row=j, column=11).value = new_filename
                            #                     j = k + 1
                            #                     # layer_file = ori_layer_file
                            #                     break
                            #                 continue
                            #             else:
                            #                 new_filename ='[WARNING索引跨级缺失]' + regularize_index_name(ws.cell(row=k, column=2).value)
                            #                 j = k + 1
                            #                 ws.cell(row=j, column=11).value = new_filename
                            #                 # layer_file = ori_layer_file
                            #                 break
                            #         else:
                            #             new_filename=ws.cell(row=k-1, column=2).value
                            #             new_filename = regularize_index_name(new_filename)
                            #             j = k
                            #             ws.cell(row=j, column=11).value = new_filename
                            #             # layer_file = ori_layer_file
                            #             break
                            #     else:
                            #         #判断自动调节级别参数，写的脑壳疼，想到后面可能讲的脑壳疼就不想写了，算了吧
                            #         # if auto_layer:
                            #         #     layer_file=layers_now_index
                            #         #     # sign_temp_layer=1
                            #         #     new_filename = '[WARNING索引自动调整]' + ws.cell(row=k-1, column=2).value
                            #         #     j = k
                            #         #     break
                            #         # else:
                            #         #     new_filename = '[WARNING层级隔页之间文件]' + ws.cell(row=k - 1, column=2).value
                            #         #     j = k
                            #         #     layer_file = ori_layer_file
                            #         #     break
                            #         new_filename = '[WARNING层级隔页之间文件]' + regularize_index_name(ws.cell(row=k - 1, column=2).value)
                            #         j = k
                            #         ws.cell(row=j, column=11).value = new_filename
                            #         # layer_file = ori_layer_file
                            #         break
                            k = j + 1
                            if k>=endlinefile + 1:
                                new_filename = ws.cell(row=j, column=2).value
                                new_filename = regularize_index_name(new_filename)
                                ws.cell(row=j, column=11).value = new_filename
                                break
                            sign_k_over = 0
                            while j + 1 <= k < endlinefile + 1:
                                k_pre = k - 1
                                k_next = k
                                while not ws.cell(row=k_pre, column=8).value:
                                    k_pre -= 1

                                while not ws.cell(row=k_next, column=8).value:
                                    if k_next < endlinefile + 1:
                                        k_next += 1
                                    else:
                                        new_filename = ws.cell(row=k_pre, column=2).value
                                        new_filename = regularize_index_name(new_filename)
                                        ws.cell(row=j, column=11).value = new_filename
                                        j = endlinefile + 1
                                        sign_k_over = 1
                                        break
                                if sign_k_over:
                                    break
                                pre_n_page = int(ws.cell(row=k_pre, column=3).value)
                                next_n_page = int(ws.cell(row=k_next, column=3).value)
                                # if next_n_page==(pre_n_page+1):
                                if next_n_page == (pre_n_page + k_next - k_pre):
                                # if next_n_page == (pre_n_page + k_next - k_pre-1):
                                    # 判断下一行是否更深
                                    layers_next_index = int(ws.cell(row=k_next, column=8).value)
                                    layers_pre_index = int(ws.cell(row=k_pre, column=8).value)
                                    if layers_next_index > layers_pre_index:
                                        # 此时应该继续深入，如果已经到末级了，就直接获取文件名
                                        if layers_next_index == layer_file:
                                            new_filename = ws.cell(row=k_next, column=2).value
                                            new_filename = regularize_index_name(new_filename)
                                            ws.cell(row=j, column=11).value = new_filename
                                            j = k_next + 1
                                            # layer_file = ori_layer_file
                                            break
                                        elif layers_next_index < layer_file:
                                            if k_next == endlinefile:
                                                new_filename = ws.cell(row=k_next, column=2).value
                                                new_filename = regularize_index_name(new_filename)
                                                ws.cell(row=j, column=11).value = new_filename
                                                j = k_next + 1
                                                # layer_file = ori_layer_file
                                                break
                                            k = k_next + 1
                                            continue
                                        else:
                                            new_filename = '[WARNING索引跨级缺失]' + regularize_index_name(
                                                ws.cell(row=k_next, column=2).value)
                                            ws.cell(row=j, column=11).value = new_filename
                                            j = k_next + 1
                                            # layer_file = ori_layer_file
                                            break
                                    else:
                                        new_filename = ws.cell(row=k_pre, column=2).value
                                        new_filename = regularize_index_name(new_filename)
                                        ws.cell(row=j, column=11).value = new_filename
                                        j = k_next
                                        # layer_file = ori_layer_file
                                        break
                                else:
                                    # new_filename = '[WARNING层级隔页之间文件]' + regularize_index_name(ws.cell(row=k_pre, column=2).value)
                                    # ws.cell(row=j, column=11).value = new_filename
                                    # 不连续的情况是有正常情况的，当kpre是文件名时，大概率不连续
                                    layers_next_index = int(ws.cell(row=k_next, column=8).value)
                                    layers_pre_index = int(ws.cell(row=k_pre, column=8).value)
                                    if layers_next_index <= layers_pre_index:
                                        new_filename = ws.cell(row=k_pre, column=2).value
                                        new_filename = regularize_index_name(new_filename)
                                        ws.cell(row=j, column=11).value = new_filename
                                    else:
                                        new_filename = '[WARNING层级隔页之间文件]' + regularize_index_name(
                                            ws.cell(row=k_pre, column=2).value)
                                        ws.cell(row=j, column=11).value = new_filename
                                    j = k_next
                                    # layer_file = ori_layer_file
                                    break
                                    # #判断自动调节级别参数
                                    # if auto_layer:
                                    #     layer_file=layers_now_index
                                    #     # sign_temp_layer=1
                                    #     new_filename = '[WARNING索引自动调整]' + ws.cell(row=k-1, column=2).value
                                    #     j = k
                                    #     break
                                    # else:
                                    #     new_filename = '[WARNING层级隔页之间文件]' + ws.cell(row=k - 1, column=2).value
                                    #     j = k
                                    #     # layer_file = ori_layer_file
                                    #     break
                            # wb.save(tb_path)
                        else:
                            new_filename=ws.cell(row=j, column=2).value
                            new_filename = regularize_index_name(new_filename)
                            ws.cell(row=j, column=11).value = new_filename
                            j = j + 1
                            # layer_file = ori_layer_file
                            continue
                    else:
                        j = j + 1
                        continue
            sign_continue_read_line = 1 #文件处理完成，继续读表
    wb.save(tb_path)

    #     if filename_now != filename_pre:
    #         filename_pre = filename_now
    #         if int(ws.cell(row=i, column=3).value) != 1:
    #             ws.insert_rows(i)
    #             ws.cell(row=i, column=1).value = ws.cell(row=i + 1, column=1).value
    #             ws.cell(row=i, column=3).value = str(1)
    #             i += 1
    #             continue
    #     ocrtxt = str(ws.cell(row=i, column=2).value)
    #     list_index = find_index(ocrtxt)
    #     if list_index[0]:
    #         ws.cell(row=i, column=6).value = '是'
    #         ws.cell(row=i, column=7).value = list_index[1]
    #         ws.cell(row=i, column=8).value = list_index[2]
    #     else:
    #         ws.cell(row=i, column=6).value = '否'
    #         ws.cell(row=i, column=7).value = list_index[1]
    #         ws.cell(row=i, column=8).value = list_index[2]
    #     i += 1
    # # print(find_index('现金收付交易'))
    # wb.save(tb_path)



if __name__ == '__main__':
    gen_ori_ocr_excel()
    tb_path=r'E:\20210917\下载完\新建文件夹\ori_tb_OCR.xlsx'
    fill_tb_index(tb_path)
    # print(find_index('第二十一章'))
    find_parent_index(tb_path)
    compute_new_filename_new_loc(tb_path)

