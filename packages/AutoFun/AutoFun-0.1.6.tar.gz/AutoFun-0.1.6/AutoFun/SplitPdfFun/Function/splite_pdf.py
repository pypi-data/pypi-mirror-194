try:
    from pdf_utils import MyPDFHandler,PDFHandleMode as mode
    from sqlite_to_excel import find_index,regularize_index_name
except:
    from .pdf_utils import MyPDFHandler,PDFHandleMode as mode
    from .sqlite_to_excel import find_index,regularize_index_name
import sys

import pandas as pd
# from Gui.Auto.processor.sqlite_to_excel import find_ilegal_character,change_num_ArabToChinese
import re
from openpyxl import load_workbook
import os
from PyPDF2 import PdfFileReader, PdfFileWriter
import traceback
import gc
#所有的索引均为数字加杠加数字形式，杠可以为零，可以为第几章的形式，其他类型索引不支持。
pd.set_option('display.max_columns',None)
# 显示所有列
pd.set_option('display.max_rows',None)
# 设置数据的显示长度，默认为50
pd.set_option('max_colwidth',200)
# 禁止自动换行(设置为Flase不自动换行，True反之)
pd.set_option('expand_frame_repr', False)

def split_pdf(old, new,first_A, first_B=None,droplist=[]):
    pdf_output = PdfFileWriter()
    pdf_file=open(old, 'rb')
    pdf_input = PdfFileReader(pdf_file)

    # 获取 pdf 共用多少页
    page_count = pdf_input.getNumPages()
    # print(page_count)
    # if not first_B:
    #     first_B=page_count

    # 将 pdf 第五页之后的页面，输出到一个新的文件
    ispdfsign=0
    if first_B:
        for i in range(first_A-1, first_B):
            if i+1 in droplist:
                continue
            pdf_output.addPage(pdf_input.getPage(i))
            ispdfsign = 1
            # print(i)
    else:
        for i in range(first_A-1, page_count):
            if i+1 in droplist:
                continue
            pdf_output.addPage(pdf_input.getPage(i))
            ispdfsign = 1
            # print(i)
    pdf_output.setPageLayout('/SinglePage')
    if new[-4:] == '.pdf':
        if ispdfsign:
            while os.path.isfile(new):
                file_name = new[:-4]
                # print('------------------------------------------------------------------')
                # print(file_name)
                file_name= file_name+'_重复'
                new =file_name+'.pdf'
            pdf_output.write(open(new, 'wb'))
        else:
            root,file=os.path.split(new)
            filename=file[0:-4]
            while os.path.isdir(os.path.join(root,filename)):
                filename+='_重复'
            os.mkdir(os.path.join(root,filename))
            while os.path.isfile(os.path.join(root,filename+'.txt')):
                filename+='_重复'
            with open(os.path.join(root,filename+'.txt'), "w") as f:
                f.write(file[0:-4]+'\n')
            new=os.path.join(root,filename+'.txt')
    else:
        print('新文件路径不是pdf文件')
    # pdf_output.write(open(new, 'wb'))
    pdf_file.close()
    return new

def add_bookmarks_layers_to_single_cutPDF(path_cut_pdf,bookmark_parent_list,):
    pdf_handler = MyPDFHandler(path_cut_pdf, mode=mode.NEWLY)
    PRTS = {}
    parent = None
    oldparent = None
    parent_default = None
    n=0
    print(bookmark_parent_list)
    for bookmarkmsg in bookmark_parent_list:
        flag_MK = 0
        key_now = re.match('(\d+(?:-\d+)*)', bookmarkmsg[0]).group(1)#改成非空符号则可以进行标签层级处理
        index_fill_table_bkmk=key_now
        key = key_now
        while flag_MK == 0:
            n_now = key.count('-')
            if PRTS.__contains__(key):
                oldparent = pdf_handler.add_one_bookmark(bookmarkmsg [0], bookmarkmsg [1], PRTS[key])
                # fill_tb_with_index(index_fill_table_bkmk, r'v', Num_of_BKMKs, dict_rowNum_index,ws)
                PRT = {key_now: oldparent}
                PRTS.update(PRT)
                # print('here')
                flag_MK = 1
            elif n_now>0:
                # print(key + '无指定目录，尝试上一级目录')
                key = re.match('((\d+(?:-\d+)*))-\d+', key).group(1)
                # print(key)
            else:
                oldparent = pdf_handler.add_one_bookmark(bookmarkmsg [0], bookmarkmsg [1])
                # fill_tb_with_index(index_fill_table_bkmk, r'v', Num_of_BKMKs, dict_rowNum_index,ws)
                PRT = {key_now: oldparent}
                PRTS.update(PRT)
                # print('here')
                flag_MK = 1
    # print('加标签保存了吗')
    #     wb.save(tb_path)
    pdf_handler.save2file(path_cut_pdf)
    return

def split_pdfs_in_tb(tb_path,scr_pdfs,dst_pdfs,iscontaincover):
    wb = load_workbook(tb_path)
    ws = wb['Sheet1']
    i = 2
    num_null = 0
    filename_pre = ''
    startlinefile = 0
    endlinefile = 0
    sign_continue_read_line = 1
    cursor_read = i
    # ori_layer_file = layer_file
    while True:
        # 逐行读，确定区间
        filename_now = ws.cell(row=i, column=1).value
        # if ws.cell(row=i, column=1).value:
        #     filename_now = ws.cell(row=i, column=1).value

        if i == 2:
            filename_pre = filename_now

        if sign_continue_read_line:
            # i=cursor_read
            if i == cursor_read:
                startlinefile = i
            if filename_now != filename_pre:
                filename_pre = filename_now
                endlinefile = i - 1
                cursor_read = i
                sign_continue_read_line = 0
                continue
            if num_null > 50:
                print('读表结束')
                break  # 读取五十个空行后自动结束
            if not ws.cell(row=i, column=1).value:
                num_null += 1
                i += 1
                continue
            i += 1
        else:
            # 在找寻好的区间内处理文件
            print('--------------------------')
            print(startlinefile)
            print(endlinefile)
            droplist=[]

            now_src_file_name = ws.cell(row=startlinefile, column=1).value
            j = startlinefile
            while startlinefile<=j<endlinefile+1:
                singlecoverpage = 0
                now_new_file_name = ws.cell(row=j, column=11).value
                # now_file_index = ws.cell(row=j, column=7).value
                # now_n_page = int(ws.cell(row=j, column=3).value)

                bookmark_parent_list=[]#包含书签文本，页码，parent，索引结构
                parent_default=None#在书签文本不含索引结构时默认为前一个有索引结构文件的子书签

                if now_new_file_name:
                    if not iscontaincover:
                        droplist.append(int(ws.cell(row=j, column=3).value))
                        singlecoverpage+=1
                    start_new_file_page=int(ws.cell(row=j, column=3).value)
                    txt_bookmarks = regularize_index_name(ws.cell(row=j, column=2).value)
                    n_page_bookmarks = int(ws.cell(row=j, column=3).value)-start_new_file_page+1
                    if not iscontaincover:
                        n_page_bookmarks = int(ws.cell(row=j, column=3).value)-start_new_file_page+1-singlecoverpage+1
                    index_bookmarks = ws.cell(row=j, column=7).value

                    bookmark_parent_list.append([txt_bookmarks, n_page_bookmarks, '', index_bookmarks])
                    if j==endlinefile:
                        if not iscontaincover:
                            droplist.append(int(ws.cell(row=j, column=3).value))
                            singlecoverpage += 1

                        end_new_file_page = ''

                        txt_bookmarks = regularize_index_name(ws.cell(row=j, column=2).value)
                        # n_page_bookmarks = int(ws.cell(row=j, column=3).value)-start_new_file_page+1-singlecoverpage
                        n_page_bookmarks = int(ws.cell(row=j, column=3).value) - start_new_file_page + 1
                        if not iscontaincover:
                            n_page_bookmarks = int(ws.cell(row=j, column=3).value) - start_new_file_page + 1 - singlecoverpage + 1
                        index_bookmarks = ws.cell(row=j, column=7).value

                        bookmark_parent_list.append([txt_bookmarks, n_page_bookmarks, '', index_bookmarks])
                        j = endlinefile + 1
                        # break
                    else:
                        for k in range(j+1,endlinefile+1):
                            if ws.cell(row=k, column=11).value:
                                end_new_file_page=int(ws.cell(row=k, column=3).value)-1
                                j=k
                                break
                            if not iscontaincover:
                                droplist.append(int(ws.cell(row=k, column=3).value))
                                singlecoverpage += 1
                            txt_bookmarks = regularize_index_name(ws.cell(row=k, column=2).value)
                            n_page_bookmarks = int(ws.cell(row=k, column=3).value)-start_new_file_page+1
                            if not iscontaincover:
                                n_page_bookmarks = int(ws.cell(row=k, column=3).value) - start_new_file_page + 1 - singlecoverpage+1
                            index_bookmarks = ws.cell(row=k, column=7).value
                            bookmark_parent_list.append([txt_bookmarks, n_page_bookmarks, '', index_bookmarks])
                            if k==endlinefile:
                                end_new_file_page = ''
                                j=endlinefile+1
                                break
                # print(1111111)
                print('$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$')
                print(now_new_file_name)
                print(start_new_file_page)
                print(end_new_file_page)
                print(bookmark_parent_list)

                # if filename_now:
                if now_new_file_name:
                    src_pdf_path=os.path.join(scr_pdfs,now_src_file_name+'.pdf')
                    new_pdf_path=os.path.join(dst_pdfs,now_new_file_name+'.pdf')
                    print("这是报错信号")
                    print(src_pdf_path,new_pdf_path,start_new_file_page,end_new_file_page,droplist)
                    new_pdf_path=split_pdf(src_pdf_path,new_pdf_path,start_new_file_page,end_new_file_page,droplist)
                    dict_parent_layer={0:None}
                    now_layer_file_index=find_index(now_new_file_name)[2]
                    if now_layer_file_index=='':
                        now_layer_file_index=0
                    save_sign=0
                    if len(bookmark_parent_list):
                        # index_list_bookmark_parent=[x[2] for x in bookmark_parent_list]
                        print('a')
                        if new_pdf_path[-4:]=='.pdf':
                            print('a1')
                            try:
                                pdf_handler = MyPDFHandler(new_pdf_path, mode=mode.NEWLY)
                            except :
                                traceback.print_exc()
                            print('b')
                            for each_bookmark_parent in bookmark_parent_list:
                                # if not now_file_index:
                                #     now_file_index=''
                                if each_bookmark_parent[3]:
                                    # if each_bookmark_parent[3].count('-')<now_file_index.count('-'):
                                    if each_bookmark_parent[3].count('-')+1<now_layer_file_index:
                                        dict_parent_layer.update({each_bookmark_parent[3].count('-')+1:None})
                                    elif each_bookmark_parent[3].count('-')+1==now_layer_file_index:
                                        if now_layer_file_index>0:
                                            dict_parent_layer.update({now_layer_file_index: None})
                                        else:
                                            parent_default= pdf_handler.add_one_bookmark(each_bookmark_parent[0], int(each_bookmark_parent[1]))
                                            save_sign=1
                                            # each_bookmark_parent[2]=parent_default
                                            dict_parent_layer.update({each_bookmark_parent[3].count('-')+1: parent_default})
                                    else:
                                        n_layer_bookmark=each_bookmark_parent[3].count('-')+1
                                        parent_layer=n_layer_bookmark-1
                                        while True:
                                            if dict_parent_layer.__contains__(parent_layer):
                                                parent_index=dict_parent_layer[parent_layer]
                                                parent_default = pdf_handler.add_one_bookmark(each_bookmark_parent[0],int(each_bookmark_parent[1]),parent_index)
                                                save_sign = 1
                                                # each_bookmark_parent[2]=parent_default
                                                dict_parent_layer.update({each_bookmark_parent[3].count('-')+1: parent_default})
                                                break
                                            elif parent_layer>0:
                                                parent_layer-=1
                                else:
                                    pdf_handler.add_one_bookmark(each_bookmark_parent[0], int(each_bookmark_parent[1]),parent_default)
                                    save_sign = 1
                            print('c')
                            if save_sign==1:
                                pdf_handler.save2file(new_pdf_path)
                                # pdf_handler.close()
                                # del pdf_handler
                                # gc.collect()

                            print('$$$$$$$$$$$$$QQQQQQQQQQQQQQ$$$$$$$$$$$$$$$$$$$')
                        elif new_pdf_path[-4:]=='.txt':
                            writesign=0
                            # if writesign==0:
                            #     with open(new_pdf_path, "a") as f:
                            #         f.write('文件页数为0，无法添加pdf书签，具体文件未处理书签内容情况如下：' + '\n')
                            for each_bookmark_parent in bookmark_parent_list:
                                # with open(new_pdf_path, "a") as f:
                                #     f.write(each_bookmark_parent[0]+'\n')
                                if each_bookmark_parent[3]:
                                    # if each_bookmark_parent[3].count('-')<now_file_index.count('-'):
                                    if each_bookmark_parent[3].count('-') + 1 < now_layer_file_index:
                                        dict_parent_layer.update({each_bookmark_parent[3].count('-') + 1: None})
                                    elif each_bookmark_parent[3].count('-') + 1 == now_layer_file_index:
                                        if now_layer_file_index > 0:
                                            dict_parent_layer.update({now_layer_file_index: None})
                                        else:
                                            if writesign == 0:
                                                with open(new_pdf_path, "a") as f:
                                                    f.write('文件页数为0，无法添加pdf书签，具体文件未处理书签内容情况如下：' + '\n')
                                            with open(new_pdf_path, "a") as f:
                                                f.write(each_bookmark_parent[0] + '\n')
                                            # parent_default = pdf_handler.add_one_bookmark(each_bookmark_parent[0],
                                            #                                               int(each_bookmark_parent[
                                            #                                                       1]))
                                            writesign = 1
                                            # each_bookmark_parent[2]=parent_default
                                            dict_parent_layer.update(
                                                {each_bookmark_parent[3].count('-') + 1: parent_default})
                                    else:
                                        n_layer_bookmark = each_bookmark_parent[3].count('-') + 1
                                        parent_layer = n_layer_bookmark - 1
                                        while True:
                                            if dict_parent_layer.__contains__(parent_layer):
                                                parent_index = dict_parent_layer[parent_layer]
                                                # parent_default = pdf_handler.add_one_bookmark(
                                                #     each_bookmark_parent[0], int(each_bookmark_parent[1]),
                                                #     parent_index)
                                                if writesign == 0:
                                                    with open(new_pdf_path, "a") as f:
                                                        f.write('文件页数为0，无法添加pdf书签，具体文件未处理书签内容情况如下：' + '\n')
                                                with open(new_pdf_path, "a") as f:
                                                    f.write(each_bookmark_parent[0] + '\n')
                                                writesign = 1
                                                # each_bookmark_parent[2]=parent_default
                                                dict_parent_layer.update(
                                                    {each_bookmark_parent[3].count('-') + 1: parent_default})
                                                break
                                            elif parent_layer > 0:
                                                parent_layer -= 1
                                else:
                                    if writesign == 0:
                                        with open(new_pdf_path, "a") as f:
                                            f.write('文件页数为0，无法添加pdf书签，具体文件未处理书签内容情况如下：' + '\n')
                                    with open(new_pdf_path, "a") as f:
                                        f.write(each_bookmark_parent[0] + '\n')
                                    writesign = 1

                                    # print()
            sign_continue_read_line = 1  # 文件处理完成，继续读表

if __name__ == '__main__':

    # scr_pdfs = r'C:\Users\RD-PC\Desktop\RGB_to_HSV\get_pdf_pic\src'
    # dst_pdfs = r'C:\Users\RD-PC\Desktop\RGB_to_HSV\get_pdf_pic\dst_pdfs'
    scr_pdfs = sys.argv[1]
    dst_pdfs = sys.argv[2]
    tb_path = sys.argv[3]
    iscontaincover = int(sys.argv[4])
    split_pdfs_in_tb(tb_path,scr_pdfs,dst_pdfs,iscontaincover)
